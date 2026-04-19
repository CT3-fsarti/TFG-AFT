from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


INF_SENTINEL = 999999.0
FLOAT_TOLERANCE = 1e-9
NODE_ID_PATTERN = re.compile(r"^[ocigd]\d+$", re.IGNORECASE)

INPUT_TABLES = {
    "tipos_nodo": "tblTiposDeNodo",
    "nodos": "tblNodos",
    "pesos_coste": "tblPesosCoste",
    "pesos_valor_operativo": "tblPesosValorOperativo",
    "canales": "tblCanales",
    "enlaces": "tblEnlaces",
}

OUTPUT_TABLES = {
    "red": "tblRed",
    "matriz_adyacencia": "tblMatrizDeAdyacencia",
    "matriz_costes": "tblMatrizPonderadaCostes",
    "matriz_valor_operativo": "tblMatrizPonderadaValorOperativo",
    "matriz_tradeoff": "tblMatrizTradeOff",
    "matriz_distancias_directas": "tblMatrizDistanciasDirectas",
    "matriz_distancias_minimas": "tblMatrizDistanciasMinimas",
    "metricas_nodos": "tblMetricasNodos",
    "metricas_red": "tblMetricasNodos20",
}

REQUIRED_COLUMNS = {
    "tipos_nodo": {"tipo", "descripcion", "capa", "notas"},
    "nodos": {"nodoid", "nombre", "tipo", "activo"},
    "pesos_coste": {"valor", "peso"},
    "pesos_valor_operativo": {
        "valor",
        "frecuencia_de_uso",
        "opacidad",
        "trazabilidad",
        "velocidad",
        "escalabilidad",
    },
    "canales": {
        "canal",
        "frecuencia_de_uso",
        "coste",
        "opacidad",
        "trazabilidad",
        "velocidad",
        "escalabilidad",
    },
    "enlaces": {"nodo_inicial", "nodo_final", "canal", "activo"},
}

NETWORK_METRIC_NOTES = {
    "distancia_media_al_destino_final_d1": {
        "metrica": "Distancia media al destino final D1",
        "que_mide": "La friccion media necesaria para llegar al destino final desde los nodos que si pueden alcanzarlo.",
        "interpretacion": "alto -> llegar a D1 exige mas friccion media\nbajo -> la topologia canaliza mejor el flujo hacia D1",
    },
    "distancia_media_total_de_la_red": {
        "metrica": "Distancia media total de la red",
        "que_mide": "La friccion media entre todos los pares de nodos conectados.",
        "interpretacion": "alto -> red mas costosa o dispersa\nbajo -> red globalmente mas eficiente",
    },
    "cercania_armonica_media_de_la_red": {
        "metrica": "Cercania armonica media de la red",
        "que_mide": "El nivel medio de proximidad estructural de los nodos dentro de la topologia.",
        "interpretacion": "alto -> red mas accesible y estructuralmente eficiente\nbajo -> red mas dispersa o fragmentada",
    },
    "eficiencia_global_de_la_red": {
        "metrica": "Eficiencia global de la red",
        "que_mide": "La capacidad global de la red para conectar nodos con baja friccion.",
        "interpretacion": "alto -> red mas eficiente globalmente\nbajo -> red menos eficiente o con mas desconexion",
    },
    "centralizacion_de_la_red": {
        "metrica": "Centralizacion de la red",
        "que_mide": "Mide el grado en que la importancia estructural de la red esta concentrada en uno o pocos nodos, a partir de la distribucion del grado ponderado total.",
        "interpretacion": "valor alto -> red mas centralizada, jerarquica o dependiente de pocos nodos\nvalor bajo -> red mas distribuida, equilibrada o descentralizada",
    },
}


class WorkbookSchemaError(ValueError):
    pass


@dataclass
class ComparisonSummary:
    name: str
    expected_rows: int
    actual_rows: int
    extra_rows: int
    missing_rows: int
    numeric_columns: list[str]
    text_columns: list[str]
    numeric_mismatches: int
    text_mismatches: int
    max_abs_diff: float

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "expected_rows": self.expected_rows,
            "actual_rows": self.actual_rows,
            "extra_rows": self.extra_rows,
            "missing_rows": self.missing_rows,
            "numeric_columns": self.numeric_columns,
            "text_columns": self.text_columns,
            "numeric_mismatches": self.numeric_mismatches,
            "text_mismatches": self.text_mismatches,
            "max_abs_diff": self.max_abs_diff,
        }


def normalize_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = "" if value is None else str(value)
    text = text.strip()
    text = text.replace("→", " -> ")
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.casefold()


def canonical_name(value: Any) -> str:
    text = normalize_text(value)
    if NODE_ID_PATTERN.fullmatch(text):
        return text.upper()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def normalize_category(value: Any) -> str:
    text = normalize_text(value)
    text = text.replace(" - ", "-").replace(" ", "")
    return text


def is_blank(value: Any) -> bool:
    return value is None or pd.isna(value) or str(value).strip() == ""


def safe_float(value: Any) -> float:
    if value is None or value == "" or pd.isna(value):
        return 0.0
    if isinstance(value, (int, float, np.number)):
        return float(value)
    return float(str(value).strip())


def canonicalize_frame(frame: pd.DataFrame) -> pd.DataFrame:
    renamed = frame.copy()
    renamed.columns = [canonical_name(column) for column in frame.columns]
    return renamed


def load_table_frame(workbook: Workbook, table_name: str) -> pd.DataFrame:
    normalized_target = canonical_name(table_name)
    for worksheet in workbook.worksheets:
        for existing_name in worksheet.tables.keys():
            table = worksheet.tables[existing_name]
            if canonical_name(existing_name) != normalized_target:
                continue
            values = [[cell.value for cell in row] for row in worksheet[table.ref]]
            if not values:
                return pd.DataFrame()
            return pd.DataFrame(values[1:], columns=values[0])
    raise KeyError(table_name)


def enforce_required_columns(frame: pd.DataFrame, alias: str) -> None:
    missing = sorted(REQUIRED_COLUMNS[alias] - set(frame.columns))
    if missing:
        raise WorkbookSchemaError(
            f"La tabla '{alias}' no contiene las columnas requeridas: {', '.join(missing)}"
        )


def prepare_base_tables(frames: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    prepared = {alias: canonicalize_frame(frame) for alias, frame in frames.items()}
    for alias in INPUT_TABLES:
        enforce_required_columns(prepared[alias], alias)

    tipos = prepared["tipos_nodo"].copy()
    tipos = tipos[~tipos["tipo"].map(is_blank)].reset_index(drop=True)

    nodos = prepared["nodos"].copy()
    nodos = nodos[~nodos["nodoid"].map(is_blank)].reset_index(drop=True)
    nodos["nodoid"] = nodos["nodoid"].astype(str).str.strip()
    nodos["activo"] = nodos["activo"].fillna(0).astype(int)

    pesos_coste = prepared["pesos_coste"].copy()
    pesos_coste = pesos_coste[~pesos_coste["valor"].map(is_blank)].reset_index(drop=True)
    pesos_coste["peso"] = pesos_coste["peso"].map(safe_float)

    pesos_valor = prepared["pesos_valor_operativo"].copy()
    pesos_valor = pesos_valor[~pesos_valor["valor"].map(is_blank)].reset_index(drop=True)
    for column in [
        "frecuencia_de_uso",
        "opacidad",
        "trazabilidad",
        "velocidad",
        "escalabilidad",
    ]:
        pesos_valor[column] = pesos_valor[column].map(safe_float)

    canales = prepared["canales"].copy()
    canales = canales[~canales["canal"].map(is_blank)].reset_index(drop=True)

    enlaces = prepared["enlaces"].copy()
    enlaces = enlaces[
        (~enlaces["nodo_inicial"].map(is_blank))
        & (~enlaces["nodo_final"].map(is_blank))
        & (~enlaces["canal"].map(is_blank))
    ].reset_index(drop=True)
    enlaces["nodo_inicial"] = enlaces["nodo_inicial"].astype(str).str.strip()
    enlaces["nodo_final"] = enlaces["nodo_final"].astype(str).str.strip()
    enlaces["canal"] = enlaces["canal"].astype(str).str.strip()
    enlaces["activo"] = enlaces["activo"].fillna(0).astype(int)

    return {
        "tipos_nodo": tipos,
        "nodos": nodos,
        "pesos_coste": pesos_coste,
        "pesos_valor_operativo": pesos_valor,
        "canales": canales,
        "enlaces": enlaces,
    }


def build_enlaces_detalle(base: dict[str, pd.DataFrame]) -> pd.DataFrame:
    nodos = base["nodos"]
    canales = base["canales"]
    pesos_coste = base["pesos_coste"]
    pesos_valor = base["pesos_valor_operativo"]
    enlaces = base["enlaces"].copy()

    nombre_por_nodo = nodos.set_index("nodoid")["nombre"].to_dict()
    canal_por_nombre = canales.set_index(canales["canal"].map(normalize_text))
    coste_map = {
        normalize_category(row.valor): row.peso
        for row in pesos_coste.itertuples(index=False)
    }
    valor_map = {
        normalize_category(row.valor): {
            "frecuencia_de_uso": row.frecuencia_de_uso,
            "opacidad": row.opacidad,
            "trazabilidad": row.trazabilidad,
            "velocidad": row.velocidad,
            "escalabilidad": row.escalabilidad,
        }
        for row in pesos_valor.itertuples(index=False)
    }

    seen_pairs: OrderedDict[tuple[str, str], str] = OrderedDict()
    enlace_ids: list[str] = []
    for pair in zip(enlaces["nodo_inicial"], enlaces["nodo_final"]):
        if pair not in seen_pairs:
            seen_pairs[pair] = f"E{len(seen_pairs) + 1:03d}"
        enlace_ids.append(seen_pairs[pair])
    enlaces["enlace"] = enlace_ids

    detalle_inicial = []
    detalle_final = []
    frecuencia = []
    coste = []
    opacidad = []
    trazabilidad = []
    velocidad = []
    escalabilidad = []

    for row in enlaces.itertuples(index=False):
        detalle_inicial.append(nombre_por_nodo.get(row.nodo_inicial, ""))
        detalle_final.append(nombre_por_nodo.get(row.nodo_final, ""))

        channel_key = normalize_text(row.canal)
        if channel_key not in canal_por_nombre.index:
            raise WorkbookSchemaError(f"Canal no encontrado en tblCanales: {row.canal}")
        channel = canal_por_nombre.loc[channel_key]
        if isinstance(channel, pd.DataFrame):
            raise WorkbookSchemaError(f"El canal aparece duplicado en tblCanales: {row.canal}")

        frecuencia_label = normalize_category(channel["frecuencia_de_uso"])
        coste_label = normalize_category(channel["coste"])

        # These mappings intentionally reproduce the workbook formulas exactly.
        # The Excel VLOOKUP indices are shifted, so the operational dimensions are
        # not taken from the conceptual columns one would expect.
        opacidad_label = normalize_category(channel.get("trazabilidad", ""))
        trazabilidad_label = normalize_category(channel.get("velocidad", ""))
        velocidad_label = normalize_category(channel.get("escalabilidad", ""))
        escalabilidad_label = normalize_category(channel.get("explicacion_opacidad", ""))

        frecuencia.append(valor_map.get(frecuencia_label, {}).get("frecuencia_de_uso", 0.0))
        coste.append(coste_map.get(coste_label, 0.0))
        opacidad.append(valor_map.get(opacidad_label, {}).get("opacidad", 0.0))
        trazabilidad.append(valor_map.get(trazabilidad_label, {}).get("trazabilidad", 0.0))
        velocidad.append(valor_map.get(velocidad_label, {}).get("velocidad", 0.0))
        escalabilidad.append(valor_map.get(escalabilidad_label, {}).get("escalabilidad", 0.0))

    enlaces["detalle_nodos_iniciales"] = detalle_inicial
    enlaces["detalle_nodos_finales"] = detalle_final
    enlaces["frecuencia_de_uso"] = frecuencia
    enlaces["coste"] = coste
    enlaces["opacidad"] = opacidad
    enlaces["trazabilidad"] = trazabilidad
    enlaces["velocidad"] = velocidad
    enlaces["escalabilidad"] = escalabilidad
    enlaces["valor_operativo"] = (
        enlaces["opacidad"]
        + enlaces["trazabilidad"]
        + enlaces["velocidad"]
        + enlaces["escalabilidad"]
    )

    active_frequency_total = (
        enlaces.loc[enlaces["activo"] == 1]
        .groupby("enlace")["frecuencia_de_uso"]
        .sum()
        .to_dict()
    )
    enlaces["frecuencia_de_uso_relativa"] = 0.0
    active_mask = enlaces["activo"] == 1
    denominators = enlaces.loc[active_mask, "enlace"].map(active_frequency_total)
    enlaces.loc[active_mask, "frecuencia_de_uso_relativa"] = np.where(
        denominators.to_numpy(dtype=float) > 0,
        enlaces.loc[active_mask, "frecuencia_de_uso"].to_numpy(dtype=float)
        / denominators.to_numpy(dtype=float),
        0.0,
    )

    ordered_columns = [
        "enlace",
        "nodo_inicial",
        "detalle_nodos_iniciales",
        "nodo_final",
        "detalle_nodos_finales",
        "canal",
        "activo",
        "frecuencia_de_uso",
        "frecuencia_de_uso_relativa",
        "coste",
        "opacidad",
        "trazabilidad",
        "velocidad",
        "escalabilidad",
        "valor_operativo",
    ]
    return enlaces[ordered_columns]


def build_red(base: dict[str, pd.DataFrame], enlaces_detalle: pd.DataFrame) -> pd.DataFrame:
    nodos = base["nodos"]
    node_active = nodos.set_index("nodoid")["activo"].to_dict()

    enlaces_ordenados = (
        enlaces_detalle[["enlace", "nodo_inicial", "nodo_final"]]
        .drop_duplicates(subset=["enlace"], keep="first")
        .reset_index(drop=True)
    )

    activos_por_enlace = enlaces_detalle.groupby("enlace")["activo"].sum().to_dict()
    active_rows = enlaces_detalle.loc[enlaces_detalle["activo"] == 1].copy()
    active_rows["coste_ponderado"] = (
        active_rows["coste"] * active_rows["frecuencia_de_uso_relativa"]
    )
    active_rows["valor_operativo_ponderado"] = (
        active_rows["valor_operativo"] * active_rows["frecuencia_de_uso_relativa"]
    )

    agregados = (
        active_rows.groupby("enlace", as_index=False)
        .agg(
            coste=("coste_ponderado", "sum"),
            valor_operativo=("valor_operativo_ponderado", "sum"),
        )
    )

    red = enlaces_ordenados.merge(agregados, on="enlace", how="left")
    red[["coste", "valor_operativo"]] = red[["coste", "valor_operativo"]].fillna(0.0)
    red["activo"] = red.apply(
        lambda row: int(
            node_active.get(row["nodo_inicial"], 0) == 1
            and node_active.get(row["nodo_final"], 0) == 1
            and activos_por_enlace.get(row["enlace"], 0) > 0
        ),
        axis=1,
    )
    tradeoff = np.zeros(len(red), dtype=float)
    red_coste = red["coste"].to_numpy(dtype=float)
    np.divide(
        red["valor_operativo"].to_numpy(dtype=float),
        red_coste,
        out=tradeoff,
        where=red_coste > 0,
    )
    red["trade_off_valor_operativo_coste"] = tradeoff

    return red[
        [
            "enlace",
            "nodo_inicial",
            "nodo_final",
            "activo",
            "coste",
            "valor_operativo",
            "trade_off_valor_operativo_coste",
        ]
    ]


def build_matrix_frame(
    node_order: list[str],
    node_names: dict[str, str],
    values: pd.DataFrame,
    value_column: str,
    default_value: float,
) -> pd.DataFrame:
    matrix = pd.DataFrame(default_value, index=node_order, columns=node_order, dtype=float)
    if value_column == "adjacency_marker":
        matrix.loc[:, :] = 0.0

    for row in values.itertuples(index=False):
        if row.activo != 1:
            continue
        origin = row.nodo_inicial
        target = row.nodo_final
        if origin not in matrix.index or target not in matrix.columns:
            continue
        matrix.at[origin, target] = 1.0 if value_column == "adjacency_marker" else safe_float(getattr(row, value_column))

    table = matrix.reset_index().rename(columns={"index": "nodo_orig_dest"})
    table.insert(0, "nombre_nodo", table["nodo_orig_dest"].map(node_names))
    return table


def to_numeric_matrix(table: pd.DataFrame, node_order: list[str]) -> pd.DataFrame:
    matrix = table.set_index("nodo_orig_dest")[node_order].copy()
    matrix = matrix.astype(float)
    matrix = matrix.reindex(index=node_order, columns=node_order)
    return matrix


def floyd_warshall_numeric(direct_matrix: pd.DataFrame) -> pd.DataFrame:
    node_order = list(direct_matrix.index)
    dist = direct_matrix.to_numpy(dtype=float).copy()
    size = dist.shape[0]
    for k in range(size):
        candidate = dist[:, [k]] + dist[[k], :]
        dist = np.minimum(dist, candidate)
    return pd.DataFrame(dist, index=node_order, columns=node_order)


def build_metrics_nodos(node_order: list[str], node_names: dict[str, str], tradeoff: pd.DataFrame, dist_min: pd.DataFrame) -> pd.DataFrame:
    tradeoff_np = tradeoff.to_numpy(dtype=float)
    dist_np = dist_min.to_numpy(dtype=float)
    entrada = tradeoff_np.sum(axis=0)
    salida_columns = node_order
    if "I5" in node_order:
        salida_columns = node_order[: node_order.index("I5") + 1]
    salida = tradeoff[salida_columns].to_numpy(dtype=float).sum(axis=1)
    total = entrada + salida
    total_max = total.max() if len(total) else 0.0
    normalizado = np.where(total_max > 0, total / total_max, 0.0)

    harmonic_mask = (dist_np > 0) & (dist_np < INF_SENTINEL)
    harmonic_values = np.zeros_like(dist_np, dtype=float)
    np.divide(1.0, dist_np, out=harmonic_values, where=harmonic_mask)
    harmonic = harmonic_values.sum(axis=1)

    return pd.DataFrame(
        {
            "nombre_nodo": [node_names[node_id] for node_id in node_order],
            "nodo": node_order,
            "grado_ponderado_de_entrada": entrada,
            "grado_ponderado_de_salida": salida,
            "grado_ponderado_total": total,
            "grado_normalizado": normalizado,
            "cercania_armonica": harmonic,
        }
    )


def build_metrics_red(node_order: list[str], dist_min: pd.DataFrame, metricas_nodos: pd.DataFrame) -> pd.DataFrame:
    dist_np = dist_min.to_numpy(dtype=float)
    harmonic = metricas_nodos["cercania_armonica"].to_numpy(dtype=float)
    total_strength = metricas_nodos["grado_ponderado_total"].to_numpy(dtype=float)

    if "D1" not in dist_min.columns:
        raise WorkbookSchemaError("No existe el nodo D1 en la matriz de distancias minimas")

    to_d1 = dist_min["D1"].to_numpy(dtype=float)
    valid_to_d1 = (to_d1 > 0) & (to_d1 < INF_SENTINEL)
    valid_all = (dist_np > 0) & (dist_np < INF_SENTINEL)

    distance_to_d1 = float(np.mean(to_d1[valid_to_d1])) if valid_to_d1.any() else 0.0
    distance_total = float(np.mean(dist_np[valid_all])) if valid_all.any() else 0.0
    harmonic_mean = float(np.mean(harmonic)) if len(harmonic) else 0.0

    efficiency_values = np.zeros_like(dist_np, dtype=float)
    np.divide(1.0, dist_np, out=efficiency_values, where=valid_all)
    efficiency_numerator = efficiency_values.sum()
    node_count = len(node_order)
    efficiency = (
        float(efficiency_numerator / (node_count * (node_count - 1)))
        if node_count > 1
        else 0.0
    )

    max_strength = float(total_strength.max()) if len(total_strength) else 0.0
    if node_count > 1 and max_strength > 0:
        centralization = float(
            (max_strength - total_strength).sum() / ((node_count - 1) * max_strength)
        )
    else:
        centralization = 0.0

    value_map = {
        "distancia_media_al_destino_final_d1": distance_to_d1,
        "distancia_media_total_de_la_red": distance_total,
        "cercania_armonica_media_de_la_red": harmonic_mean,
        "eficiencia_global_de_la_red": efficiency,
        "centralizacion_de_la_red": centralization,
    }

    rows = []
    for key, meta in NETWORK_METRIC_NOTES.items():
        rows.append(
            {
                "metrica": meta["metrica"],
                "valor": value_map[key],
                "que_mide": meta["que_mide"],
                "interpretacion_de_los_valores": meta["interpretacion"],
            }
        )
    return pd.DataFrame(rows)


def build_artifacts(base: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    nodos = base["nodos"]
    node_order = nodos["nodoid"].tolist()
    node_names = nodos.set_index("nodoid")["nombre"].to_dict()

    enlaces_detalle = build_enlaces_detalle(base)
    red = build_red(base, enlaces_detalle)

    matriz_adyacencia = build_matrix_frame(node_order, node_names, red, "adjacency_marker", 0.0)
    matriz_costes = build_matrix_frame(node_order, node_names, red, "coste", 0.0)
    matriz_valor = build_matrix_frame(node_order, node_names, red, "valor_operativo", 0.0)
    matriz_tradeoff = build_matrix_frame(
        node_order, node_names, red, "trade_off_valor_operativo_coste", 0.0
    )

    tradeoff_numeric = to_numeric_matrix(matriz_tradeoff, node_order)
    direct_numeric = pd.DataFrame(
        INF_SENTINEL, index=node_order, columns=node_order, dtype=float
    )
    for node_id in node_order:
        direct_numeric.at[node_id, node_id] = 0.0

    positive_mask = tradeoff_numeric.to_numpy(dtype=float) > 0
    direct_np = direct_numeric.to_numpy(dtype=float).copy()
    np.divide(
        1.0,
        tradeoff_numeric.to_numpy(dtype=float),
        out=direct_np,
        where=positive_mask,
    )
    np.fill_diagonal(direct_np, 0.0)
    direct_numeric = pd.DataFrame(direct_np, index=node_order, columns=node_order)

    matriz_distancias_directas = direct_numeric.reset_index().rename(
        columns={"index": "nodo_orig_dest"}
    )
    matriz_distancias_directas.insert(
        0, "nombre_nodo", matriz_distancias_directas["nodo_orig_dest"].map(node_names)
    )

    min_numeric = floyd_warshall_numeric(direct_numeric)
    matriz_distancias_minimas = min_numeric.reset_index().rename(
        columns={"index": "nodo_orig_dest"}
    )
    matriz_distancias_minimas.insert(
        0, "nombre_nodo", matriz_distancias_minimas["nodo_orig_dest"].map(node_names)
    )

    metricas_nodos = build_metrics_nodos(node_order, node_names, tradeoff_numeric, min_numeric)
    metricas_red = build_metrics_red(node_order, min_numeric, metricas_nodos)

    return {
        "enlaces_detalle": enlaces_detalle,
        "red": red,
        "matriz_adyacencia": matriz_adyacencia,
        "matriz_costes": matriz_costes,
        "matriz_valor_operativo": matriz_valor,
        "matriz_tradeoff": matriz_tradeoff,
        "matriz_distancias_directas": matriz_distancias_directas,
        "matriz_distancias_minimas": matriz_distancias_minimas,
        "metricas_nodos": metricas_nodos,
        "metricas_red": metricas_red,
    }


def load_expected_outputs(workbook: Workbook, node_order: list[str]) -> dict[str, pd.DataFrame]:
    frames: dict[str, pd.DataFrame] = {}
    for alias, table_name in OUTPUT_TABLES.items():
        frame = canonicalize_frame(load_table_frame(workbook, table_name))
        frames[alias] = frame

    frames["red"] = frames["red"].loc[~frames["red"]["enlace"].map(is_blank)].reset_index(drop=True)
    frames["metricas_nodos"] = frames["metricas_nodos"].loc[
        frames["metricas_nodos"]["nodo"].isin(node_order)
    ].reset_index(drop=True)
    frames["metricas_red"] = frames["metricas_red"].loc[
        ~frames["metricas_red"]["metrica"].map(is_blank)
    ].reset_index(drop=True)

    for alias in [
        "matriz_adyacencia",
        "matriz_costes",
        "matriz_valor_operativo",
        "matriz_tradeoff",
        "matriz_distancias_directas",
        "matriz_distancias_minimas",
    ]:
        frames[alias] = frames[alias].loc[
            frames[alias]["nodo_orig_dest"].isin(node_order), ["nombre_nodo", "nodo_orig_dest", *node_order]
        ].reset_index(drop=True)

    return frames


def compare_tables(
    name: str,
    expected: pd.DataFrame,
    actual: pd.DataFrame,
    key_columns: list[str],
    numeric_columns: list[str],
    text_columns: list[str] | None = None,
) -> ComparisonSummary:
    text_columns = text_columns or []
    expected_view = expected.copy()
    actual_view = actual.copy()

    for column in key_columns + text_columns:
        if column in expected_view.columns:
            expected_view[column] = expected_view[column].map(lambda value: "" if value is None else str(value))
        if column in actual_view.columns:
            actual_view[column] = actual_view[column].map(lambda value: "" if value is None else str(value))

    for column in key_columns:
        if column in expected_view.columns:
            expected_view[column] = expected_view[column].map(normalize_text)
        if column in actual_view.columns:
            actual_view[column] = actual_view[column].map(normalize_text)

    expected_view = expected_view.sort_values(key_columns).reset_index(drop=True)
    actual_view = actual_view.sort_values(key_columns).reset_index(drop=True)

    expected_keys = set(map(tuple, expected_view[key_columns].itertuples(index=False, name=None)))
    actual_keys = set(map(tuple, actual_view[key_columns].itertuples(index=False, name=None)))
    missing_keys = expected_keys - actual_keys
    extra_keys = actual_keys - expected_keys

    merged = expected_view.merge(
        actual_view,
        on=key_columns,
        how="outer",
        suffixes=("_expected", "_actual"),
        indicator=True,
    )
    overlap = merged.loc[merged["_merge"] == "both"].copy()

    numeric_mismatches = 0
    max_abs_diff = 0.0
    for column in numeric_columns:
        left = overlap[f"{column}_expected"].fillna(0.0).astype(float)
        right = overlap[f"{column}_actual"].fillna(0.0).astype(float)
        diffs = (left - right).abs()
        numeric_mismatches += int((diffs > FLOAT_TOLERANCE).sum())
        if not diffs.empty:
            max_abs_diff = max(max_abs_diff, float(diffs.max()))

    text_mismatches = 0
    for column in text_columns:
        left = overlap[f"{column}_expected"].fillna("").map(normalize_text)
        right = overlap[f"{column}_actual"].fillna("").map(normalize_text)
        text_mismatches += int((left != right).sum())

    return ComparisonSummary(
        name=name,
        expected_rows=len(expected_view),
        actual_rows=len(actual_view),
        extra_rows=len(extra_keys),
        missing_rows=len(missing_keys),
        numeric_columns=numeric_columns,
        text_columns=text_columns,
        numeric_mismatches=numeric_mismatches,
        text_mismatches=text_mismatches,
        max_abs_diff=max_abs_diff,
    )


def compare_artifacts(
    generated: dict[str, pd.DataFrame], expected: dict[str, pd.DataFrame], node_order: list[str]
) -> list[ComparisonSummary]:
    summaries = []
    summaries.append(
        compare_tables(
            "red",
            expected["red"],
            generated["red"],
            ["enlace"],
            ["activo", "coste", "valor_operativo", "trade_off_valor_operativo_coste"],
            ["nodo_inicial", "nodo_final"],
        )
    )

    for alias in [
        "matriz_adyacencia",
        "matriz_costes",
        "matriz_valor_operativo",
        "matriz_tradeoff",
        "matriz_distancias_directas",
        "matriz_distancias_minimas",
    ]:
        summaries.append(
            compare_tables(
                alias,
                expected[alias],
                generated[alias],
                ["nodo_orig_dest"],
                node_order,
                ["nombre_nodo"],
            )
        )

    summaries.append(
        compare_tables(
            "metricas_nodos",
            expected["metricas_nodos"],
            generated["metricas_nodos"],
            ["nodo"],
            [
                "grado_ponderado_de_entrada",
                "grado_ponderado_de_salida",
                "grado_ponderado_total",
                "grado_normalizado",
                "cercania_armonica",
            ],
            ["nombre_nodo"],
        )
    )
    summaries.append(
        compare_tables(
            "metricas_red",
            expected["metricas_red"],
            generated["metricas_red"],
            ["metrica"],
            ["valor"],
            ["que_mide", "interpretacion_de_los_valores"],
        )
    )
    return summaries


def export_artifacts(artifacts: dict[str, pd.DataFrame], export_dir: Path) -> None:
    export_dir.mkdir(parents=True, exist_ok=True)
    for name, frame in artifacts.items():
        frame.to_csv(export_dir / f"{name}.csv", index=False)


def print_validation_summary(base: dict[str, pd.DataFrame], generated: dict[str, pd.DataFrame]) -> None:
    print("Tablas base cargadas:")
    print(f"- nodos: {len(base['nodos'])}")
    print(f"- canales: {len(base['canales'])}")
    print(f"- enlaces detalle de entrada: {len(base['enlaces'])}")
    print(f"- enlaces detalle generados: {len(generated['enlaces_detalle'])}")
    print(f"- enlaces agregados: {len(generated['red'])}")


def print_comparison_summary(summaries: list[ComparisonSummary]) -> None:
    print("\nComparacion con las hojas calculadas del Excel:")
    for summary in summaries:
        status = "OK"
        if (
            summary.missing_rows
            or summary.extra_rows
            or summary.numeric_mismatches
            or summary.text_mismatches
        ):
            status = "DIFF"
        print(
            f"- {summary.name}: {status} | rows expected={summary.expected_rows} actual={summary.actual_rows} "
            f"missing={summary.missing_rows} extra={summary.extra_rows} "
            f"numeric_mismatches={summary.numeric_mismatches} text_mismatches={summary.text_mismatches} "
            f"max_abs_diff={summary.max_abs_diff:.12g}"
        )


def build_bridge(workbook_path: Path) -> tuple[dict[str, pd.DataFrame], dict[str, pd.DataFrame], list[str]]:
    workbook_values = load_workbook(workbook_path, data_only=True, keep_vba=True)
    input_frames = {
        alias: load_table_frame(workbook_values, table_name)
        for alias, table_name in INPUT_TABLES.items()
    }
    base = prepare_base_tables(input_frames)
    artifacts = build_artifacts(base)
    node_order = base["nodos"]["nodoid"].tolist()
    return base, artifacts, node_order


def run_cli(args: argparse.Namespace) -> int:
    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {workbook_path}")

    base, artifacts, node_order = build_bridge(workbook_path)
    print_validation_summary(base, artifacts)

    if args.export_dir:
        export_artifacts(artifacts, Path(args.export_dir))
        print(f"\nCSV exportados en: {Path(args.export_dir)}")

    if args.compare:
        workbook_values = load_workbook(workbook_path, data_only=True, keep_vba=True)
        expected = load_expected_outputs(workbook_values, node_order)
        summaries = compare_artifacts(artifacts, expected, node_order)
        print_comparison_summary(summaries)
        if args.json:
            json_payload = [summary.to_dict() for summary in summaries]
            print("\nJSON:")
            print(json.dumps(json_payload, indent=2, ensure_ascii=False))

    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Carga, valida y reconstruye en Python las tablas calculadas del modelo FT."
    )
    parser.add_argument(
        "--workbook",
        default="Diseño Red FT v5a.xlsm",
        help="Ruta al workbook xlsm/xlsx que se quiere analizar.",
    )
    parser.add_argument(
        "--compare",
        action="store_true",
        help="Compara las tablas reconstruidas contra las hojas calculadas del Excel.",
    )
    parser.add_argument(
        "--export-dir",
        help="Directorio opcional donde exportar las tablas generadas en CSV.",
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Muestra el resumen de comparacion tambien en JSON.",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        return run_cli(args)
    except Exception as exc:
        print(f"Error: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())