import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import networkx as nx
from pyvis.network import Network
from openpyxl import load_workbook
import os
import base64
import json

# ==========================================
# ESCUDO PROTECTOR PARA VERTEX AI
# ==========================================
try:
    from google.oauth2 import service_account
    import vertexai
    from vertexai.generative_models import GenerativeModel
    VERTEX_AVAILABLE = True
except ImportError:
    VERTEX_AVAILABLE = False

# ==========================================
# 1. CONFIGURACIÓN DE LA PÁGINA (BRANDING)
# ==========================================
st.set_page_config(page_title="Marina Sarti Pineda - Inteligencia FT/AML", layout="wide")

st.markdown("""
<style>
    :root {
        --uc3m-blue: #005691;
    }
    
    .stMarkdown h2 {
        font-size: 24px !important;
        color: #005691 !important;
        padding-bottom: 5px !important;
        font-weight: 600 !important;
        border-bottom: 2px solid #E6E9EF;
        margin-top: 20px;
    }
    
    .stMarkdown h3 {
        font-size: 18px !important;
        color: #333 !important;
        font-weight: 600 !important;
    }
    
    .centered-text { text-align: center; }
    .language-block { display: flex; align-items: flex-start; margin-bottom: 18px; }
    
    .flag-img { 
        width: 40px !important; 
        height: 26px !important; 
        object-fit: cover !important; 
        margin-right: 15px; 
        margin-top: 2px;
        border: 1px solid #ddd;
    }
    
    .flag-inline {
        height: 14px; 
        width: 22px; 
        object-fit: cover; 
        vertical-align: middle; 
        margin-right: 6px; 
        margin-bottom: 3px;
        border: 1px solid #ccc; 
        border-radius: 2px;
    }
    
    .title-text { font-size: 15px; color: #222; line-height: 1.5; }
    
    .profile-photo {
        width: 160px; height: 160px; border-radius: 50% !important; 
        object-fit: cover !important; display: block;
        margin-left: auto; margin-right: auto; border: 3px solid var(--uc3m-blue); 
    }
</style>
""", unsafe_allow_html=True)

def get_base64_of_bin_file(bin_file):
    try:
        with open(bin_file, 'rb') as f: return base64.b64encode(f.read()).decode()
    except FileNotFoundError: return None

assets_status = {
    "Combined Logo B64": get_base64_of_bin_file("Logo_uc3m_PSL.png"),
    "Flag ES B64": get_base64_of_bin_file("logo_ES.png"),
    "Flag FR B64": get_base64_of_bin_file("logo_FR.png"),
    "Flag GB B64": get_base64_of_bin_file("logo_GB.png"),
    "Marina Photo B64": get_base64_of_bin_file("marina_circular.png")
}

html_flag_es = f"<img src='data:image/png;base64,{assets_status['Flag ES B64']}' class='flag-inline'>" if assets_status['Flag ES B64'] else ""
html_flag_fr = f"<img src='data:image/png;base64,{assets_status['Flag FR B64']}' class='flag-inline'>" if assets_status['Flag FR B64'] else ""
html_flag_gb = f"<img src='data:image/png;base64,{assets_status['Flag GB B64']}' class='flag-inline'>" if assets_status['Flag GB B64'] else ""

# ==========================================
# 2. ESTRUCTURA DE DISEÑO PRINCIPAL (COLUMNAS)
# ==========================================
col_main, col_profile = st.columns([3, 1])

with col_profile:
    st.markdown("""<div style="margin-top: 30px;"></div>""", unsafe_allow_html=True)
    with st.container():
        st.markdown("<h3 class='centered-text'>👤 Perfil de la Autora</h3>", unsafe_allow_html=True)
        
        if assets_status["Marina Photo B64"]:
             st.markdown(f"""<div style="margin-bottom: 15px;"><img src='data:image/png;base64,{assets_status["Marina Photo B64"]}' class='profile-photo' /></div>""", unsafe_allow_html=True)
        
        st.markdown(f"<p class='centered-text' style='font-size: 18px; margin-bottom: 5px;'><strong>Marina Sarti Pineda</strong></p>", unsafe_allow_html=True)
        
        st.markdown("""
        <div style="text-align: center; margin-bottom: 15px;">
            <a href="https://www.linkedin.com/in/marina-sarti-pineda-27211b29b/?originalSubdomain=es" target="_blank" style="text-decoration: none; font-size: 14px; color: #005691;">🔗 Perfil de LinkedIn</a>
        </div>
        """, unsafe_allow_html=True)
        
        # STACK TECNOLÓGICO
        st.markdown("""
        <hr style="margin: 10px 0 15px 0; border-top: 1px solid #E6E9EF;">
        <p class='centered-text' style='font-size: 15px; color: #333; margin-bottom: 10px;'><strong>🛠️ Stack Tecnológico</strong></p>
        <div style="display: flex; flex-wrap: wrap; gap: 6px; justify-content: center; margin-bottom: 20px;">
            <a href="https://www.python.org/" target="_blank" style="text-decoration: none; background-color: #F0F4F8; color: #005691; padding: 3px 10px; border-radius: 12px; font-size: 12.5px; border: 1px solid #D9E2EC; font-weight: 500;">Python</a>
            <a href="https://networkx.org/" target="_blank" style="text-decoration: none; background-color: #F0F4F8; color: #005691; padding: 3px 10px; border-radius: 12px; font-size: 12.5px; border: 1px solid #D9E2EC; font-weight: 500;">NetworkX</a>
            <a href="https://pandas.pydata.org/" target="_blank" style="text-decoration: none; background-color: #F0F4F8; color: #005691; padding: 3px 10px; border-radius: 12px; font-size: 12.5px; border: 1px solid #D9E2EC; font-weight: 500;">Pandas</a>
            <a href="https://streamlit.io/" target="_blank" style="text-decoration: none; background-color: #F0F4F8; color: #005691; padding: 3px 10px; border-radius: 12px; font-size: 12.5px; border: 1px solid #D9E2EC; font-weight: 500;">Streamlit</a>
            <a href="https://pyvis.readthedocs.io/" target="_blank" style="text-decoration: none; background-color: #F0F4F8; color: #005691; padding: 3px 10px; border-radius: 12px; font-size: 12.5px; border: 1px solid #D9E2EC; font-weight: 500;">PyVis</a>
            <a href="https://cloud.google.com/vertex-ai" target="_blank" style="text-decoration: none; background-color: #F0F4F8; color: #005691; padding: 3px 10px; border-radius: 12px; font-size: 12.5px; border: 1px solid #D9E2EC; font-weight: 500;">Vertex AI</a>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown(f"""
        <div style="font-size: 14px; line-height: 1.5; color: #333; text-align: justify;">
            <p style="margin-bottom: 8px;">{html_flag_es} Graduada en Economía por la <strong>Universidad Carlos III de Madrid y París Dauphine (DTI)</strong>, Marina combina el rigor analítico macroeconómico con una vocación por la seguridad global, la inteligencia financiera y la tecnología.</p>
            <hr style="margin: 12px 0; border-top: 1px solid #eee;">
            <p style="margin-bottom: 8px;">{html_flag_fr} Diplômée en Économie de l'<strong>Universidad Carlos III de Madrid et Paris Dauphine (DTI)</strong>, Marina allie la rigueur analytique macroéconomique à une vocation pour la sécurité mondiale, le renseignement financier et la technologie.</p>
            <hr style="margin: 12px 0; border-top: 1px solid #eee;">
            <p>{html_flag_gb} Graduated in Economics from <strong>Universidad Carlos III de Madrid and Paris Dauphine (DTI)</strong>, Marina combines macroeconomic analytical rigor with a vocation for global security, financial intelligence, and technology.</p>
        </div>
        """, unsafe_allow_html=True)

with col_main:
    col_logo, col_titles = st.columns([1, 2.5])
    with col_logo:
        if assets_status["Combined Logo B64"]:
             st.markdown(f"""<img src='data:image/png;base64,{assets_status["Combined Logo B64"]}' style='width: 100%; height: auto; margin-top: 10px;' />""", unsafe_allow_html=True)
    with col_titles:
        st.markdown("""
        <div style="text-align: left; padding-top: 10px;">
            <div style="margin-bottom: 6px;">
                <span style="font-style: italic; color: #555; font-size: 18px;">Trabajo Fin de Grado - </span>
                <strong style="color: #005691; font-size: 24px;">Doble Titulación Internacional en Economía.</strong>
            </div>
            <div style="color: #222; font-style: italic; font-weight: bold; font-size: 18px; line-height: 1.4;">
                Redes de Financiación del Terrorismo:<br>
                Simulación y análisis mediante Economía de Redes y Teoría de Juegos
            </div>
        </div>
        """, unsafe_allow_html=True)
        
    st.markdown("<hr style='margin-top: 15px; margin-bottom: 20px; border-top: 1px solid #111;'>", unsafe_allow_html=True)
    
    st.markdown("## 🎓 Sinopsis del Proyecto")
    with st.expander("Leer Sinopsis / Abstract Completo", expanded=False):
        tab_es, tab_en, tab_fr = st.tabs(["Español", "English", "Français"])
        with tab_es: st.markdown("<div style='font-size: 15px; text-align: justify; color: #333;'>La Financiación del Terrorismo (FT) consiste en la captación de fondos, lo cual abarca el proceso de solicitud, recaudación, provisión y puesta a disposición de dinero o activos con el fin de facilitar o potenciar la capacidad de cualquier persona u organización para llevar a cabo actividades relacionadas con el terrorismo. En el caso concreto de España, la Ley 10/2010 establece un marco riguroso frente al suministro, depósito o distribución de fondos.<br><br>Tanto los grandes grupos organizados como las pequeñas células e incluso los actores individuales necesitan dinero para el desarrollo de la actividad terrorista. La literatura académica y los informes institucionales coinciden en que la falta de fondos limita drásticamente su capacidad operativa, siendo la FT un elemento vertebrador del terrorismo global.<br><br>Este trabajo fundamenta su análisis en información reciente, utilizando ejemplos observados en los últimos años que resultan representativos de las dinámicas contemporáneas. El objetivo principal consiste en construir una <strong>simulación de red de financiación del terrorismo</strong> basada en la evidencia recogida en la literatura especializada (tipologías del GAFI, EBA, etc.).<br><br>Sobre dicha simulación se realiza un análisis estructural mediante herramientas propias del análisis de redes (Economía de Redes) y la Teoría de Juegos, incluyendo el estudio de métricas de centralidad, la importancia relativa de los nodos clave (chokepoints) y la resiliencia del sistema ante intervenciones policiales. El modelo resultante constituye una representación realista y fundamentada empíricamente, derivando en un modelo analítico altamente útil para la inteligencia financiera y el diseño de políticas de seguridad.</div>", unsafe_allow_html=True)
        with tab_en: st.markdown("<div style='font-size: 15px; text-align: justify; color: #333;'>Terrorist Financing (TF) involves the raising of funds, which encompasses the process of soliciting, collecting, providing, and making available money or assets to facilitate or enhance the capacity of any individual or organization to carry out terrorist activities. In Spain, Law 10/2010 establishes a rigorous framework against the supply, deposit, or distribution of funds.<br><br>Large organized groups, small cells, and lone actors require money to carry out terrorist activities. Academic literature and institutional reports agree that a lack of funds drastically limits their operational capacity, making TF a structural backbone of global terrorism.<br><br>This paper bases its analysis on recent information, using examples observed in recent years that are representative of contemporary dynamics. The main objective is to build a <strong>simulation of a terrorist financing network</strong> based on evidence gathered from specialized literature (FATF typologies, EBA, etc.).<br><br>A structural analysis is performed on this simulation using Network Economics and Game Theory, including the study of centrality metrics, the relative importance of key nodes (chokepoints), and the system's resilience to law enforcement interventions. The resulting model constitutes a realistic and empirically grounded representation, resulting in an analytical model highly useful for financial intelligence and security policy design.</div>", unsafe_allow_html=True)
        with tab_fr: st.markdown("<div style='font-size: 15px; text-align: justify; color: #333;'>Le Financement du Terrorisme (FT) consiste en la collecte de fonds, ce qui englobe le processus de sollicitation, de rassemblement, de fourniture et de mise à disposition d'argent ou d'actifs dans le but de faciliter la capacité à mener des activités terroristes. En Espagne, la loi 10/2010 établit un cadre rigoureux contre la fourniture ou la distribution de fonds.<br><br>Les grands groupes organisés, les petites cellules et les acteurs individuels ont besoin d'argent pour développer leurs activités terroristes. La littérature académique s'accorde à dire que le manque de fonds limite considérablement leur capacité opérationnelle, le FT étant un élément structurant du terrorisme mondial.<br><br>Ce travail fonde son analyse sur des informations récentes, en utilisant des exemples représentatifs des dynamiques contemporaines. L'objectif principal est de construire une <strong>simulation d'un réseau de financement du terrorisme</strong> basée sur les preuves recueillies dans la littérature spécialisée (typologies du GAFI, ABE, etc.).<br><br>Une analyse structurelle est réalisée sur cette simulation à l'aide de l'Économie des Réseaux et de la Théorie des Jeux, incluant l'étude des métriques de centralité, l'importance des nœuds clés (points d'étranglement) et la résilience du système face aux interventions policières. Le modèle qui en résulte constitue une représentation réaliste et empiriquement fondée, aboutissant à un outil analytique très utile pour le renseignement financier.</div>", unsafe_allow_html=True)

# ==========================================
# 3. FUNCIONES DE CARGA Y ESTILADO
# ==========================================
def leer_tabla_excel(wb, nombre_tabla_buscada):
    for hoja in wb.worksheets:
        if nombre_tabla_buscada in hoja.tables:
            rango = hoja.tables[nombre_tabla_buscada].ref
            datos_celdas = hoja[rango]
            filas = [[celda.value for celda in fila] for fila in datos_celdas]
            return pd.DataFrame(filas[1:], columns=filas[0])
    return pd.DataFrame() 

def aplicar_estilos(df_in):
    df_safe = df_in.copy()
    df_safe.columns = df_safe.columns.astype(str)
    df_safe.loc[:, ~df_safe.columns.duplicated()]

    styler = df_safe.style.set_properties(**{'text-align': 'center'})
    styler = styler.set_table_styles([
        dict(selector='th', props=[('text-align', 'center')]),
        dict(selector='td', props=[('text-align', 'center')])
    ])
    
    styler = styler.format(lambda v: f"{v:.0f}" if isinstance(v, (int, float)) and pd.notna(v) else v)
    
    if 'Activo' in df_safe.columns:
        def color_activo(val):
            try:
                v = int(val)
                if v == 1: return 'background-color: #C6EFCE; color: #006100; font-weight: bold;'
                elif v == 0: return 'background-color: #FFC7CE; color: #9C0006; font-weight: bold;'
            except: pass
            return ''
        if hasattr(styler, 'map'): styler = styler.map(color_activo, subset=['Activo'])
        else: styler = styler.applymap(color_activo, subset=['Activo'])
            
    return styler

def aplicar_estilo_matriz(df_in, decimales=0):
    df_safe = df_in.copy()
    df_safe.index = df_safe.index.astype(str)
    df_safe.columns = df_safe.columns.astype(str)
    df_safe.loc[~df_safe.index.duplicated(keep='first')]
    df_safe.loc[:, ~df_safe.columns.duplicated()]

    styler = df_safe.style.set_properties(**{'text-align': 'center'})
    styler = styler.set_table_styles([
        dict(selector='th', props=[('text-align', 'center')]),
        dict(selector='td', props=[('text-align', 'center')])
    ])
    
    def formato_celda(v):
        if pd.isna(v) or v == 0 or v == "0" or v == "": return ""
        try: return f"{float(v):.{decimales}f}"
        except: return str(v)
        
    return styler.format(formato_celda)

# ==========================================
# 4. FLUJO DE DATOS
# ==========================================
st.markdown("---")
st.markdown("<div style='text-align: center; color: #555; font-size: 14px;'>FLUJO DEL SIMULADOR</div>", unsafe_allow_html=True)

archivo_subido = st.file_uploader("Sube tu archivo Excel alternativo (Opcional)", type=["xlsx"])

wb = None
if archivo_subido is not None:
    wb = load_workbook(archivo_subido, data_only=True)
elif os.path.exists("Diseño Red FT.xlsx"):
    wb = load_workbook("Diseño Red FT.xlsx", data_only=True)
else:
    st.info("💡 Sube un archivo Excel para comenzar.")

if wb is not None:
    df_tipos = leer_tabla_excel(wb, "tblTiposDeNodo")
    df_nodos_original = leer_tabla_excel(wb, "tblNodos")
    df_enlaces_original = leer_tabla_excel(wb, "tblEnlaces")
    df_pesos = leer_tabla_excel(wb, "tblPesos")
    df_pond_costes = leer_tabla_excel(wb, "tblMatrizPonderadaCostes")
    # Atención a la capitalización exacta de la tabla en el Excel:
    df_pond_valor = leer_tabla_excel(wb, "tblMatrizPonderadaValoroperativo")
    df_tradeoff = leer_tabla_excel(wb, "tblMatrizTradeOff")
    df_distancias = leer_tabla_excel(wb, "tblMatrizDistancias")
    df_metricas = leer_tabla_excel(wb, "tblGradoPonderado")

    # ==========================================
    # FASE 1: BASE DE DATOS Y PARAMETRIZACIÓN
    # ==========================================
    st.markdown("## ⚙️ Fase 1: Base de Datos y Parametrización")
    tab_sim1, tab_sim2, tab_sim3, tab_sim4 = st.tabs(["🟢 Nodos (Actores)", "🔗 Enlaces (Rutas)", "🏷️ Tipos de Nodo", "⚖️ Matriz de Pesos (Metodología)"])

    with tab_sim1:
        cols_nodos = [c for c in ['Activo', 'NodoID', 'Nombre', 'Tipo', 'Descripción'] if c in df_nodos_original.columns]
        if not cols_nodos: cols_nodos = df_nodos_original.columns.tolist()
        df_n_sub = df_nodos_original[cols_nodos]
        df_nodos_editado = st.data_editor(aplicar_estilos(df_n_sub), use_container_width=True, num_rows="dynamic", key="editor_nodos")
        
    with tab_sim2:
        cols_enlaces = [c for c in ['Activo', 'Nodo Origen', 'Nodo Destino', 'Tipo de Enlace', 'Exposición', 'Coste', 'Capacidad', 'Eficiencia'] if c in df_enlaces_original.columns]
        if not cols_enlaces: cols_enlaces = df_enlaces_original.columns.tolist()
        df_e_sub = df_enlaces_original[cols_enlaces]
        df_enlaces_editado = st.data_editor(aplicar_estilos(df_e_sub), use_container_width=True, num_rows="dynamic", key="editor_enlaces")
        
    with tab_sim3: 
        st.dataframe(aplicar_estilos(df_tipos), use_container_width=True)
        
    with tab_sim4:
        if not df_pesos.empty:
            st.dataframe(aplicar_estilos(df_pesos), use_container_width=True, hide_index=True)

    # Procesamiento interno
    df_tipos = df_tipos.dropna(how='all')
    df_nodos = df_nodos_editado.dropna(how='all')
    df_enlaces = df_enlaces_editado.dropna(subset=['Nodo Origen', 'Nodo Destino'])

    df_nodos['Activo'] = pd.to_numeric(df_nodos['Activo'], errors='coerce')
    df_enlaces['Activo'] = pd.to_numeric(df_enlaces['Activo'], errors='coerce')

    df_nodos = pd.merge(df_nodos, df_tipos[['Tipo', 'Capa']], on='Tipo', how='left')
    df_nodos['Capa'] = pd.to_numeric(df_nodos['Capa'], errors='coerce').fillna(0)

    nodos_activos = df_nodos[df_nodos['Activo'] == 1]
    enlaces_activos = df_enlaces[df_enlaces['Activo'] == 1]

    G = nx.DiGraph()
    colores_capa = {'O': '#FF9999', 'I': '#99CCFF', 'G': '#FFCC99', 'D': '#99FF99'}

    for _, row in nodos_activos.iterrows():
        tipo_nodo = str(row['Tipo']).strip()
        nombre_nodo = str(row.get('Nombre', ''))
        G.add_node(str(row['NodoID']).strip(), label=str(row['NodoID']).strip() + (" - " + nombre_nodo if nombre_nodo else ""), color=colores_capa.get(tipo_nodo, '#CCCCCC'), level=int(row['Capa']), title=f"Tipo: {str(row.get('Descripción', tipo_nodo))}")

    for _, row in enlaces_activos.iterrows():
        origen = str(row['Nodo Origen']).strip()
        destino = str(row['Nodo Destino']).strip()
        if origen in G.nodes and destino in G.nodes:
            exposicion = str(row.get('Exposición', '')).strip().lower()
            if 'medio-alto' in exposicion: color_flecha = '#90EE90'
            elif 'medio-bajo' in exposicion: color_flecha = '#FFB84D'
            elif 'alto' in exposicion: color_flecha = '#008000'
            elif 'medio' in exposicion: color_flecha = '#FFD700'
            elif 'bajo' in exposicion: color_flecha = '#FF4444'
            else: color_flecha = '#999999'
            G.add_edge(origen, destino, color=color_flecha, title=f"Exposición: {row.get('Exposición', 'N/A')} | Coste: {row.get('Coste', 'N/A')}")

    # ==========================================
    # FASE 2: TOPOLOGÍA DE LA RED
    # ==========================================
    st.markdown("## 🌐 Fase 2: Topología Interactiva de la Red")
    
    if not nodos_activos.empty:
        num_capas = pd.to_numeric(nodos_activos['Capa'], errors='coerce').nunique()
        max_nodos_por_capa = pd.to_numeric(nodos_activos['Capa'], errors='coerce').value_counts().max()
    else:
        num_capas, max_nodos_por_capa = 1, 1
        
    sep_horizontal = max(int((1440 - 200) / (num_capas - 1 if num_capas > 1 else 1)), 250)
    sep_vertical = max(int((650 - 120) / max_nodos_por_capa), 60)

    net = Network(height="650px", width="100%", directed=True, bgcolor="#ffffff", font_color="black")
    net.from_nx(G)
    # Configuración de red orgánica (Físicas activadas, sin columnas estrictas)
    net.set_options("""
    {
      "physics": {
        "enabled": true,
        "barnesHut": {
          "gravitationalConstant": -3000,
          "centralGravity": 0.2,
          "springLength": 200,
          "springConstant": 0.05,
          "damping": 0.09,
          "avoidOverlap": 0.5
        }
      },
      "edges": { 
        "smooth": { "type": "continuous" }, 
        "arrows": { "to": { "enabled": true, "scaleFactor": 0.6 } } 
      },
      "nodes": { "font": { "size": 16, "face": "Arial" } },
      "interaction": { "zoomView": true, "dragNodes": true, "hover": true }
    }
    """)

    net.save_graph("mapa_interactivo.html")

    gcol1, gcol2 = st.columns([4, 1])
    with gcol1:
        components.html(open("mapa_interactivo.html", 'r', encoding='utf-8').read(), height=670)

    with gcol2:
        st.subheader("📊 Análisis")
        st.metric("Total Nodos Activos", len(G.nodes))
        st.metric("Total Rutas Activas", len(G.edges))
        
        st.markdown("### Mapa de Calor (Exposición)")
        st.markdown("""
        <div style="line-height: 2; font-size: 15px;">
            🔴 <strong>Bajo:</strong> Canal opaco<br>
            🟠 <strong>Medio-Bajo:</strong> Riesgo latente<br>
            🟡 <strong>Medio:</strong> Neutro<br>
            🟢 <strong>Medio-Alto / Alto:</strong> Canal expuesto
        </div>
        """, unsafe_allow_html=True)

    # ==========================================
    # FASE 3: MODELADO MATEMÁTICO
    # ==========================================
    st.markdown("## 🔢 Fase 3: Modelado Matemático (Teoría de Grafos)")
    with st.expander("Matrices Matemáticas del Sistema", expanded=False):
        tab_matriz1, tab_matriz2, tab_matriz3, tab_matriz4, tab_matriz5, tab_matriz6 = st.tabs([
            "1️⃣ Adyacencia", "2️⃣ Costes", "3️⃣ Valor Operativo", "4️⃣ Trade-Off", "5️⃣ Distancias", "6️⃣ Métricas"
        ])

        with tab_matriz1:
            if len(G.nodes) > 0:
                matriz_adyacencia = nx.to_pandas_adjacency(G, dtype=int)
                st.dataframe(aplicar_estilo_matriz(matriz_adyacencia, 0), use_container_width=True)
        with tab_matriz2:
            if not df_pond_costes.empty:
                df_pond_costes.set_index(df_pond_costes.columns[0], inplace=True)
                st.dataframe(aplicar_estilo_matriz(df_pond_costes, 0), use_container_width=True)
        with tab_matriz3:
            if not df_pond_valor.empty:
                df_pond_valor.set_index(df_pond_valor.columns[0], inplace=True)
                st.dataframe(aplicar_estilo_matriz(df_pond_valor, 0), use_container_width=True)
        with tab_matriz4:
            if not df_tradeoff.empty:
                df_tradeoff.set_index(df_tradeoff.columns[0], inplace=True)
                st.dataframe(aplicar_estilo_matriz(df_tradeoff, 2), use_container_width=True)
        with tab_matriz5:
            if not df_distancias.empty:
                df_distancias.set_index(df_distancias.columns[0], inplace=True)
                st.dataframe(aplicar_estilo_matriz(df_distancias, 0), use_container_width=True)
        with tab_matriz6:
            if not df_metricas.empty:
                df_metricas.set_index(df_metricas.columns[0], inplace=True)
                st.dataframe(aplicar_estilo_matriz(df_metricas, 0), use_container_width=True)

    # ==========================================
    # FASE 4: ASISTENTE DE INTELIGENCIA (VERTEX AI)
    # ==========================================
    st.markdown("## 🤖 Fase 4: Asistente AFTIA")
    st.markdown("Consulta a AFTIA sobre vulnerabilidades, tipologías y tácticas de disrupción policial en base a los datos matemáticos y estructurales de la red.")

    if not VERTEX_AVAILABLE:
        st.error("⚠️ Falta instalar la librería de Google Cloud (`google-cloud-aiplatform`).")
    else:
        try:
            if "gcp_service_account_json" not in st.secrets:
                st.warning("⚠️ Faltan las credenciales en Streamlit Secrets.")
            else:
                credenciales_json = json.loads(st.secrets["gcp_service_account_json"], strict=False)
                credenciales_gcp = service_account.Credentials.from_service_account_info(credenciales_json)
                vertexai.init(project="aft-simulator", location="us-central1", credentials=credenciales_gcp)
                
                rol_analista = """
                Tu nombre en clave es AFTIA (Analizador de Financiación del Terrorismo mediante Inteligencia Artificial).
                ERES UN SISTEMA EXPERTO EN INTELIGENCIA FINANCIERA (PBC/FT).
                REGLA 1: Eres un analista especialista en redes de Financiación del Terrorismo. Usa terminología institucional y analítica (chokepoints, tipologías GAFI, resiliencia, nodos clave).
                REGLA 2: Responde EXCLUSIVAMENTE a preguntas sobre el análisis de la red proporcionada, tácticas de disrupción policial, prevención de blanqueo o financiación del terrorismo basándote rigurosamente en las matrices matemáticas y el grafo.
                REGLA 3: Si el usuario te pregunta algo fuera de este ámbito profesional, debes declinar cortésmente recordando tu estricta función como herramienta de inteligencia AFTIA.
                """
                model = GenerativeModel("gemini-2.5-pro")

                if "chat_history" not in st.session_state:
                    st.session_state.chat_history = []

                # ==========================================
                # EXTRACCIÓN MASIVA DE DATOS PARA AFTIA (100% DE LA RED Y MATRICES)
                # ==========================================
                
                # 1. Nodos Activos Completos
                if not nodos_activos.empty:
                    nodos_activos_str = nodos_activos.to_csv(index=False, sep='|')
                else:
                    nodos_activos_str = "No hay nodos activos."
                
                # 2. Nodos Inactivos Completos
                nodos_inactivos_df = df_nodos[df_nodos['Activo'] == 0]
                if not nodos_inactivos_df.empty:
                    nodos_inactivos_str = nodos_inactivos_df.to_csv(index=False, sep='|')
                else:
                    nodos_inactivos_str = "Ninguno. La red está operando al 100% de su capacidad original."
                    
                # 3. Enlaces Completos
                if not enlaces_activos.empty:
                    enlaces_str = enlaces_activos.to_csv(index=False, sep='|')
                else:
                    enlaces_str = "No hay rutas conectando los nodos."

                # 4. Matrices y Métricas
                metricas_str = df_metricas.to_csv(sep='|') if not df_metricas.empty else "No disponible."
                tradeoff_str = df_tradeoff.to_csv(sep='|') if not df_tradeoff.empty else "No disponible."
                costes_str = df_pond_costes.to_csv(sep='|') if not df_pond_costes.empty else "No disponible."
                valor_op_str = df_pond_valor.to_csv(sep='|') if not df_pond_valor.empty else "No disponible."
                distancias_str = df_distancias.to_csv(sep='|') if not df_distancias.empty else "No disponible."
                
                # 5. Adyacencia en Vivo
                adyacencia_str = nx.to_pandas_adjacency(G, dtype=int).to_csv(sep='|') if len(G.nodes) > 0 else "No disponible."

                contexto_red = f"""
                {rol_analista}
                
                ESTADO ACTUAL DE LA RED SIMULADA (DATOS COMPLETOS Y MATRICES MATEMÁTICAS):
                
                1. RESUMEN:
                - Nodos (actores) activos: {len(G.nodes)}
                - Rutas financieras (enlaces) activas: {len(G.edges)}
                
                2. BASE DE DATOS DE ACTORES ACTIVOS (Roles y Descripciones completas):
                {nodos_activos_str}
                
                3. ACTORES NEUTRALIZADOS (Arrestados / Desactivados por el usuario):
                {nodos_inactivos_str}
                
                4. TOPOLOGÍA DE RUTAS FINANCIERAS (Costes, Eficiencias y Capacidades operativas):
                {enlaces_str}
                
                5. MÉTRICAS DE CENTRALIDAD (Importancia matemática de cada nodo):
                {metricas_str}
                
                6. MATRIZ DE TRADE-OFF (Equilibrio de Nash, relación coste-beneficio para cada canal):
                {tradeoff_str}
                
                7. MATRIZ PONDERADA DE COSTES (Fricción operativa):
                {costes_str}
                
                8. MATRIZ PONDERADA DE VALOR OPERATIVO (Capacidad y volumen de flujo):
                {valor_op_str}
                
                9. MATRIZ DE DISTANCIAS (Saltos mínimos entre actores):
                {distancias_str}
                
                10. MATRIZ BINARIA DE ADYACENCIA (Conexiones topológicas directas):
                {adyacencia_str}
                """

                col_chat1, col_chat2 = st.columns([4, 1])
                with col_chat2:
                    if st.button("🔄 Recalcular Análisis", help="Borra la memoria y vuelve a pedir a AFTIA que analice todos los datos en vivo de la red"):
                        st.session_state.chat_history = []
                        st.rerun()

                # AUTO-PROMPT INICIAL
                if len(st.session_state.chat_history) == 0:
                    with st.spinner("AFTIA está cruzando los datos topológicos y las matrices matemáticas..."):
                        prompt_inicial = f"{contexto_red}\n\nINSTRUCCIÓN OCULTA DEL SISTEMA: Acaba de arrancar el simulador. Actúa de forma proactiva. Saluda cordialmente presentándote como AFTIA. Tienes acceso a todos los datos matemáticos y estructurales arriba. Redacta un diagnóstico analítico inicial riguroso (máximo 2 párrafos) apoyándote en las métricas de centralidad y la matriz de trade-off provistas para identificar quién es el actor clave (Chokepoint) y sugiere a la analista una primera línea de investigación."
                        response = model.generate_content(prompt_inicial)
                        st.session_state.chat_history.append({"role": "assistant", "content": response.text})

                for msg in st.session_state.chat_history:
                    st.chat_message(msg["role"]).write(msg["content"])

                user_query = st.chat_input("Haz una pregunta técnica matemática o táctica a AFTIA...")

                if user_query:
                    st.chat_message("user").write(user_query)
                    st.session_state.chat_history.append({"role": "user", "content": user_query})

                    historial_str = "\n".join([f"{'Analista' if m['role']=='user' else 'AFTIA'}: {m['content']}" for m in st.session_state.chat_history[-4:]])
                    
                    prompt_completo = f"""
                    {contexto_red}
                    
                    MEMORIA RECIENTE DE LA CONVERSACIÓN:
                    {historial_str}
                    
                    Analista: {user_query}
                    AFTIA:
                    """

                    with st.chat_message("assistant"):
                        with st.spinner("AFTIA está evaluando las matrices..."):
                            resp = model.generate_content(prompt_completo)
                            st.write(resp.text)
                            st.session_state.chat_history.append({"role": "assistant", "content": resp.text})

        except json.JSONDecodeError:
            st.error("❌ Error leyendo el archivo JSON de credenciales.")
        except Exception as e:
            st.error("❌ Ha ocurrido un error al conectar con Vertex AI.")
            st.code(str(e))