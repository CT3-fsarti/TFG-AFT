from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path
from typing import Any

import dash_bootstrap_components as dbc
import dash_cytoscape as cyto
import pandas as pd
import plotly.express as px
from dash import ALL, Dash, Input, Output, State, MATCH, dash_table, dcc, html, no_update
from dash.exceptions import PreventUpdate
from openpyxl import load_workbook

from ft_excel_bridge import (
    INF_SENTINEL,
    INPUT_TABLES,
    WorkbookSchemaError,
    build_artifacts,
    load_table_frame,
    prepare_base_tables,
)


cyto.load_extra_layouts()

APP_TITLE = "Comparador FT en Dash"
DEFAULT_WORKBOOKS = ["Diseño Red FT v5a.xlsm", "Diseño Red FT.xlsx"]
TOPOLOGY_DEFINITIONS = [
    {
        "id": "meta_red",
        "label": "Meta-Red",
        "description": "Escenario base del workbook. Actua como referencia fija para comparar el resto de topologias.",
        "editable": False,
    },
    {
        "id": "estrella",
        "label": "Estrella",
        "description": "Activa o desactiva nodos y enlaces para quedarte con una topologia centralizada.",
        "editable": True,
    },
    {
        "id": "anillo",
        "label": "Anillo",
        "description": "Usa la Meta-Red como base y deja activas solo las rutas que formen un circuito.",
        "editable": True,
    },
    {
        "id": "multi_hub",
        "label": "Multi-Hub",
        "description": "Construye una red con varios nodos concentradores comparables entre si.",
        "editable": True,
    },
]
METRIC_COLUMN_ORDER = [
    "Distancia media al destino final D1",
    "Distancia media total de la red",
    "Cercania armonica media de la red",
    "Eficiencia global de la red",
    "Centralizacion de la red",
]
ARTIFACT_TABS = [
    ("red", "Red agregada", False),
    ("matriz_adyacencia", "Adyacencia", True),
    ("matriz_costes", "Costes", True),
    ("matriz_valor_operativo", "Valor operativo", True),
    ("matriz_tradeoff", "Trade-Off", True),
    ("matriz_distancias_minimas", "Distancias minimas", True),
    ("metricas_nodos", "Metricas por nodo", False),
    ("metricas_red", "Metricas de red", False),
]


def find_default_workbook_path() -> Path | None:
    for candidate in DEFAULT_WORKBOOKS:
        path = Path(candidate)
        if path.exists():
            return path
    return None


def load_base_model_from_bytes(workbook_bytes: bytes) -> dict[str, pd.DataFrame]:
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True, keep_vba=True)
    frames = {
        alias: load_table_frame(workbook, table_name)
        for alias, table_name in INPUT_TABLES.items()
    }
    return prepare_base_tables(frames)


def serialize_frames(frames: dict[str, pd.DataFrame]) -> dict[str, list[dict[str, Any]]]:
    return {name: frame.to_dict("records") for name, frame in frames.items()}


def deserialize_frames(payload: dict[str, list[dict[str, Any]]]) -> dict[str, pd.DataFrame]:
    return {name: pd.DataFrame(records) for name, records in payload.items()}


def copy_frames(frames: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    return {name: frame.copy(deep=True) for name, frame in frames.items()}


APP_STATE: dict[str, Any] = {
    "version": 0,
    "workbook_source": None,
    "workbook_bytes": None,
    "base_model": None,
    "palette": None,
    "scenario_states": {},
    "error": None,
}


def initialize_scenario_states(base_model: dict[str, pd.DataFrame]) -> dict[str, dict[str, list[dict[str, Any]]]]:
    nodes_editor = build_nodes_editor_frame(base_model).to_dict("records")
    links_editor = build_links_editor_frame(base_model).to_dict("records")
    states: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for scenario in TOPOLOGY_DEFINITIONS:
        states[scenario["id"]] = {
            "nodes": [dict(row) for row in nodes_editor],
            "links": [dict(row) for row in links_editor],
        }
    return states


def set_app_workbook(workbook_bytes: bytes, workbook_source: str) -> int:
    try:
        base_model = load_base_model_from_bytes(workbook_bytes)
        palette = extract_excel_palette(workbook_bytes)
        APP_STATE["version"] = int(APP_STATE.get("version", 0)) + 1
        APP_STATE["workbook_source"] = workbook_source
        APP_STATE["workbook_bytes"] = workbook_bytes
        APP_STATE["base_model"] = base_model
        APP_STATE["palette"] = palette
        APP_STATE["scenario_states"] = initialize_scenario_states(base_model)
        APP_STATE["error"] = None
        return APP_STATE["version"]
    except Exception as exc:
        APP_STATE["error"] = str(exc)
        raise


def initialize_app_state() -> None:
    workbook_path = find_default_workbook_path()
    if workbook_path is None:
        APP_STATE["error"] = (
            "No se encontro ningun workbook por defecto. Deja 'Diseño Red FT v5a.xlsm' en la carpeta del proyecto o sube uno manualmente."
        )
        return
    set_app_workbook(workbook_path.read_bytes(), workbook_path.name)


def normalize_active(value: Any) -> int:
    if value in (True, "True", "true", 1, "1"):
        return 1
    return 0


def build_nodes_editor_frame(base_model: dict[str, pd.DataFrame]) -> pd.DataFrame:
    capas = base_model["tipos_nodo"][["tipo", "descripcion", "capa"]].drop_duplicates()
    nodos = base_model["nodos"].merge(capas, on="tipo", how="left")
    editor = nodos[["nodoid", "nombre", "tipo", "descripcion", "capa", "activo"]].copy()
    editor = editor.rename(
        columns={
            "nodoid": "NodoID",
            "nombre": "Nombre",
            "tipo": "Tipo",
            "descripcion": "Descripcion",
            "capa": "Capa",
            "activo": "Activo",
        }
    )
    editor["Activo"] = editor["Activo"].map(normalize_active)
    return editor


def build_links_editor_frame(base_model: dict[str, pd.DataFrame]) -> pd.DataFrame:
    nombres = base_model["nodos"].set_index("nodoid")["nombre"].to_dict()
    enlaces = base_model["enlaces"][["nodo_inicial", "nodo_final", "canal", "activo"]].copy()
    enlaces.insert(0, "row_id", enlaces.index.astype(int))
    enlaces["nombre_origen"] = enlaces["nodo_inicial"].map(nombres)
    enlaces["nombre_destino"] = enlaces["nodo_final"].map(nombres)
    editor = enlaces[
        [
            "row_id",
            "nodo_inicial",
            "nombre_origen",
            "nodo_final",
            "nombre_destino",
            "canal",
            "activo",
        ]
    ].copy()
    editor = editor.rename(
        columns={
            "row_id": "row_id",
            "nodo_inicial": "Nodo Inicial",
            "nombre_origen": "Nombre Origen",
            "nodo_final": "Nodo Final",
            "nombre_destino": "Nombre Destino",
            "canal": "Canal",
            "activo": "Activo",
        }
    )
    editor["Activo"] = editor["Activo"].map(normalize_active)
    return editor


def apply_scenario_actives(
    base_model: dict[str, pd.DataFrame],
    nodes_records: list[dict[str, Any]],
    links_records: list[dict[str, Any]],
) -> dict[str, pd.DataFrame]:
    scenario = copy_frames(base_model)
    node_actives = {
        str(row.get("NodoID", "")).strip(): normalize_active(row.get("Activo"))
        for row in nodes_records
    }
    link_actives = {
        int(row.get("row_id")): normalize_active(row.get("Activo"))
        for row in links_records
        if row.get("row_id") is not None
    }
    scenario["nodos"]["activo"] = scenario["nodos"]["nodoid"].map(node_actives).fillna(0).astype(int)
    scenario["enlaces"]["activo"] = (
        scenario["enlaces"].index.to_series().map(link_actives).fillna(0).astype(int)
    )
    return scenario


def locate_table(workbook, table_name: str):
    for worksheet in workbook.worksheets:
        if table_name in worksheet.tables:
            return worksheet, worksheet.tables[table_name]
    raise KeyError(table_name)


def color_to_css(color_obj: Any) -> str | None:
    if color_obj is None:
        return None
    try:
        rgb = color_obj.rgb
    except Exception:
        return None
    if not rgb or rgb == "00000000":
        return None
    if len(rgb) == 8:
        return f"#{rgb[2:]}"
    if len(rgb) == 6:
        return f"#{rgb}"
    return None


def cell_fill_css(cell) -> str | None:
    return color_to_css(cell.fill.fgColor) or color_to_css(cell.fill.start_color)


def cell_font_css(cell) -> str | None:
    return color_to_css(cell.font.color)


def extract_excel_palette(workbook_bytes: bytes) -> dict[str, str]:
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False, keep_vba=True)
    ws_nodos, tbl_nodos = locate_table(workbook, "tblNodos")
    ws_costes, tbl_costes = locate_table(workbook, "tblMatrizPonderadaCostes")
    nodos_cells = ws_nodos[tbl_nodos.ref]
    costes_cells = ws_costes[tbl_costes.ref]

    header_cell = nodos_cells[0][0]
    matrix_index_cell = costes_cells[1][1]
    matrix_value_cell = costes_cells[1][2]

    return {
        "header_bg": cell_fill_css(header_cell) or "#1F4E78",
        "header_fg": cell_font_css(header_cell) or "#FFFFFF",
        "table_bg": "#FFFFFF",
        "table_fg": "#1B263B",
        "sheet_bg": "#F5F1E8",
        "card_bg": "#FFFCF7",
        "card_border": "#D7C9B6",
        "accent": "#1F4E78",
        "muted": "#5C6B73",
        "matrix_index_bg": cell_fill_css(matrix_index_cell) or "#1F4E78",
        "matrix_index_fg": cell_font_css(matrix_index_cell) or "#FFFFFF",
        "matrix_value_bg": cell_fill_css(matrix_value_cell) or "#F8D7DA",
        "matrix_value_fg": cell_font_css(matrix_value_cell) or "#000000",
        "active_on_bg": "#C6EFCE",
        "active_on_fg": "#006100",
        "active_off_bg": "#FFC7CE",
        "active_off_fg": "#9C0006",
    }


def sanitize_frame_for_dash(frame: pd.DataFrame, is_matrix: bool = False) -> pd.DataFrame:
    clean = frame.copy()
    clean.columns = [str(column) for column in clean.columns]
    for column in clean.columns:
        if pd.api.types.is_numeric_dtype(clean[column]):
            clean[column] = clean[column].astype(float)
            clean.loc[clean[column] >= INF_SENTINEL, column] = None
            if is_matrix:
                clean[column] = clean[column].round(3)
            else:
                clean[column] = clean[column].round(4)
        clean[column] = clean[column].where(pd.notna(clean[column]), None)
    return clean


def build_table_columns(frame: pd.DataFrame, editable: bool) -> list[dict[str, Any]]:
    columns = []
    for column in frame.columns:
        column_def: dict[str, Any] = {
            "name": column,
            "id": column,
            "editable": editable and column == "Activo",
        }
        if column == "Activo" and editable:
            column_def["presentation"] = "dropdown"
        columns.append(column_def)
    return columns


def build_table_style_conditions(
    frame: pd.DataFrame,
    palette: dict[str, str],
    table_kind: str,
) -> list[dict[str, Any]]:
    conditions: list[dict[str, Any]] = []
    if "Activo" in frame.columns:
        conditions.extend(
            [
                {
                    "if": {"column_id": "Activo", "filter_query": "{Activo} = 1"},
                    "backgroundColor": palette["active_on_bg"],
                    "color": palette["active_on_fg"],
                    "fontWeight": 700,
                },
                {
                    "if": {"column_id": "Activo", "filter_query": "{Activo} = 0"},
                    "backgroundColor": palette["active_off_bg"],
                    "color": palette["active_off_fg"],
                    "fontWeight": 700,
                },
            ]
        )

    if table_kind == "matrix":
        label_columns = list(frame.columns[:2])
        if len(label_columns) > 1:
            conditions.append(
                {
                    "if": {"column_id": label_columns[1]},
                    "backgroundColor": palette["matrix_index_bg"],
                    "color": palette["matrix_index_fg"],
                    "fontWeight": 700,
                }
            )
        if label_columns:
            conditions.append(
                {
                    "if": {"column_id": label_columns[0]},
                    "fontWeight": 600,
                    "backgroundColor": "#FBF7F0",
                }
            )
        for column in frame.columns[2:]:
            if pd.api.types.is_numeric_dtype(frame[column]):
                conditions.append(
                    {
                        "if": {"column_id": column, "filter_query": f"{{{column}}} > 0"},
                        "backgroundColor": palette["matrix_value_bg"],
                        "color": palette["matrix_value_fg"],
                    }
                )

    if table_kind == "metrics":
        first_column = frame.columns[0]
        conditions.append(
            {
                "if": {"column_id": first_column},
                "fontWeight": 700,
                "backgroundColor": "#FBF7F0",
            }
        )

    return conditions


def build_dash_table(
    frame: pd.DataFrame,
    palette: dict[str, str],
    table_id: Any,
    table_kind: str,
    editable: bool = False,
    hidden_columns: list[str] | None = None,
) -> dash_table.DataTable:
    clean = sanitize_frame_for_dash(frame, is_matrix=table_kind == "matrix")
    page_size = 10 if len(clean) > 10 else max(len(clean), 1)
    table = dash_table.DataTable(
        id=table_id,
        data=clean.to_dict("records"),
        columns=build_table_columns(clean, editable=editable),
        hidden_columns=hidden_columns or [],
        editable=editable,
        dropdown={
            "Activo": {
                "options": [
                    {"label": "1", "value": 1},
                    {"label": "0", "value": 0},
                ]
            }
        }
        if editable and "Activo" in clean.columns
        else {},
        page_size=page_size,
        style_table={
            "overflowX": "auto",
            "border": f"1px solid {palette['card_border']}",
            "borderRadius": "12px",
            "width": "100%",
        },
        style_header={
            "backgroundColor": palette["header_bg"],
            "color": palette["header_fg"],
            "fontWeight": 700,
            "textAlign": "center",
            "border": f"1px solid {palette['card_border']}",
            "whiteSpace": "normal",
            "height": "auto",
        },
        style_cell={
            "backgroundColor": palette["table_bg"],
            "color": palette["table_fg"],
            "textAlign": "center",
            "padding": "8px 10px",
            "whiteSpace": "normal",
            "height": "auto",
            "fontFamily": "Segoe UI, sans-serif",
            "fontSize": "13px",
            "border": "1px solid #E4DBCF",
        },
        style_data_conditional=build_table_style_conditions(clean, palette, table_kind),
        fill_width=True,
        cell_selectable=False,
    )
    return table


def build_topology_graph_elements(base_model: dict[str, pd.DataFrame], red_frame: pd.DataFrame) -> list[dict[str, Any]]:
    tipos = base_model["tipos_nodo"][["tipo", "descripcion", "capa"]].drop_duplicates()
    nodos = base_model["nodos"].merge(tipos, on="tipo", how="left")
    nodos_activos = nodos[nodos["activo"] == 1].copy()
    enlaces_activos = red_frame[red_frame["activo"] == 1].copy()

    nodos_activos["capa"] = pd.to_numeric(nodos_activos["capa"], errors="coerce").fillna(0).astype(int)
    nodos_activos = nodos_activos.sort_values(["capa", "nodoid"]).reset_index(drop=True)
    positions: dict[str, dict[str, float]] = {}
    layer_spacing = 320
    base_y_spacing = 120
    for capa, layer_frame in nodos_activos.groupby("capa", sort=True):
        total = len(layer_frame)
        offset = (total - 1) / 2
        for index, row in enumerate(layer_frame.itertuples(index=False)):
            positions[str(row.nodoid).strip()] = {
                "x": float(capa * layer_spacing),
                "y": float((index - offset) * base_y_spacing),
            }

    elements: list[dict[str, Any]] = []
    for row in nodos_activos.itertuples(index=False):
        node_id = str(row.nodoid).strip()
        node_name = str(row.nombre).strip() if pd.notna(row.nombre) else ""
        elements.append(
            {
                "data": {
                    "id": node_id,
                    "label": node_id,
                    "nombre": node_name,
                    "tipo": str(row.tipo).strip(),
                    "descripcion": str(row.descripcion).strip() if pd.notna(row.descripcion) else "",
                    "capa": int(row.capa),
                },
                "classes": f"tipo-{str(row.tipo).strip().lower()}",
                "position": positions.get(node_id, {"x": 0.0, "y": 0.0}),
            }
        )

    for row in enlaces_activos.itertuples(index=False):
        tradeoff = float(row.trade_off_valor_operativo_coste) if pd.notna(row.trade_off_valor_operativo_coste) else 0.0
        if tradeoff >= 3:
            edge_class = "edge-high"
        elif tradeoff >= 2:
            edge_class = "edge-medium"
        else:
            edge_class = "edge-low"
        elements.append(
            {
                "data": {
                    "source": str(row.nodo_inicial).strip(),
                    "target": str(row.nodo_final).strip(),
                    "label": "",
                    "coste": f"{float(row.coste):.3f}",
                    "valor": f"{float(row.valor_operativo):.3f}",
                    "tradeoff": f"{tradeoff:.3f}",
                },
                "classes": edge_class,
            }
        )
    return elements


def build_cytoscape_stylesheet(palette: dict[str, str]) -> list[dict[str, Any]]:
    return [
        {
            "selector": "node",
            "style": {
                "label": "data(label)",
                "font-size": "13px",
                "font-weight": 700,
                "font-family": "Segoe UI",
                "text-wrap": "none",
                "text-valign": "center",
                "text-halign": "center",
                "color": "#14213D",
                "background-color": "#BFC8D6",
                "border-width": 2,
                "border-color": "#FFFFFF",
                "width": 56,
                "height": 56,
                "overlay-opacity": 0,
                "text-outline-width": 0,
            },
        },
        {"selector": ".tipo-o", "style": {"background-color": "#D66D75"}},
        {"selector": ".tipo-c", "style": {"background-color": "#E8A66B"}},
        {"selector": ".tipo-i", "style": {"background-color": "#5DA9E9"}},
        {"selector": ".tipo-g", "style": {"background-color": "#7BC47F"}},
        {"selector": ".tipo-d", "style": {"background-color": "#6C5CE7", "color": "#FFFFFF"}},
        {
            "selector": "node:selected",
            "style": {
                "border-width": 4,
                "border-color": palette["accent"],
                "z-index": 999,
            },
        },
        {
            "selector": "edge",
            "style": {
                "curve-style": "bezier",
                "target-arrow-shape": "triangle",
                "arrow-scale": 1.1,
                "width": 3,
                "line-color": "#9AA5B1",
                "target-arrow-color": "#9AA5B1",
                "opacity": 0.8,
                "font-size": "9px",
                "label": "data(label)",
                "text-background-opacity": 0,
                "text-background-color": "#FFFCF7",
                "text-background-padding": "2px",
                "overlay-opacity": 0,
            },
        },
        {
            "selector": ".edge-high",
            "style": {"line-color": "#1B9E77", "target-arrow-color": "#1B9E77", "width": 3.5},
        },
        {
            "selector": ".edge-medium",
            "style": {"line-color": "#E6AB02", "target-arrow-color": "#E6AB02", "width": 3},
        },
        {
            "selector": ".edge-low",
            "style": {"line-color": "#D95F02", "target-arrow-color": "#D95F02", "width": 2.5},
        },
        {
            "selector": "edge:selected",
            "style": {
                "line-color": palette["accent"],
                "target-arrow-color": palette["accent"],
                "width": 4.5,
                "z-index": 999,
            },
        },
    ]


def build_graph_info_panel(palette: dict[str, str], node_data: dict[str, Any] | None = None, edge_data: dict[str, Any] | None = None):
    title = "Detalle del grafo"
    rows: list[Any]
    if node_data:
        title = f"Nodo {node_data.get('id', '')}"
        rows = [
            html.Div([html.Strong("Nombre: "), html.Span(node_data.get("nombre") or "-")]),
            html.Div([html.Strong("Tipo: "), html.Span(node_data.get("tipo") or "-")]),
            html.Div([html.Strong("Capa: "), html.Span(str(node_data.get("capa", "-")))]),
            html.Div([html.Strong("Descripcion: "), html.Span(node_data.get("descripcion") or "-")]),
        ]
    elif edge_data:
        title = f"Enlace {edge_data.get('source', '')} -> {edge_data.get('target', '')}"
        rows = [
            html.Div([html.Strong("Coste: "), html.Span(edge_data.get("coste") or "-")]),
            html.Div([html.Strong("Valor operativo: "), html.Span(edge_data.get("valor") or "-")]),
            html.Div([html.Strong("Trade-off: "), html.Span(edge_data.get("tradeoff") or "-")]),
        ]
    else:
        rows = [
            html.Div("El grafo usa una disposicion jerarquica izquierda-derecha basada en las capas del modelo."),
            html.Div("Pulsa sobre un nodo o enlace para ver su detalle aqui."),
            html.Hr(),
            html.Div([html.Strong("Colores de nodo: "), html.Span("O, C, I, G y D representan capas funcionales distintas.")]),
            html.Div([html.Strong("Colores de enlace: "), html.Span("Verde alto trade-off, amarillo medio, naranja bajo.")]),
        ]

    return dbc.Card(
        dbc.CardBody(
            [
                html.Div(title, style={"fontWeight": 700, "color": palette["accent"], "marginBottom": "10px"}),
                html.Div(rows, className="small", style={"lineHeight": "1.7"}),
            ]
        ),
        style={
            "background": palette["card_bg"],
            "border": f"1px solid {palette['card_border']}",
            "borderRadius": "16px",
            "boxShadow": "0 10px 24px rgba(31, 78, 120, 0.08)",
            "marginBottom": "12px",
        },
    )


def build_metric_cards(metricas_red: pd.DataFrame, palette: dict[str, str]) -> dbc.Row:
    values = metricas_red[["metrica", "valor"]].copy()
    values["valor"] = pd.to_numeric(values["valor"], errors="coerce").fillna(0.0)
    value_map = dict(zip(values["metrica"], values["valor"]))
    cards = []
    for metric_name in METRIC_COLUMN_ORDER:
        cards.append(
            dbc.Col(
                dbc.Card(
                    dbc.CardBody(
                        [
                            html.Div(metric_name, className="small text-muted", style={"minHeight": "44px"}),
                            html.Div(f"{value_map.get(metric_name, 0.0):.4f}", style={"fontSize": "1.35rem", "fontWeight": 700, "color": palette["accent"]}),
                        ]
                    ),
                    style={
                        "background": palette["card_bg"],
                        "border": f"1px solid {palette['card_border']}",
                        "borderRadius": "16px",
                        "boxShadow": "0 10px 24px rgba(31, 78, 120, 0.08)",
                    },
                ),
                md=12,
                lg=12,
                className="mb-3",
            )
        )
    return dbc.Row(cards, className="g-2")


def build_artifact_tabs(artifacts: dict[str, pd.DataFrame], palette: dict[str, str], scenario_id: str) -> dcc.Tabs:
    tabs = []
    for artifact_name, label, is_matrix in ARTIFACT_TABS:
        frame = artifacts[artifact_name].copy()
        table_kind = "matrix" if is_matrix else ("metrics" if artifact_name.startswith("metricas") else "standard")
        hidden_columns = ["row_id"] if "row_id" in frame.columns else []
        tabs.append(
            dcc.Tab(
                label=label,
                value=f"{scenario_id}-{artifact_name}",
                children=html.Div(
                    build_dash_table(
                        frame,
                        palette,
                        table_id={"type": "artifact-table", "scenario": scenario_id, "artifact": artifact_name},
                        table_kind=table_kind,
                        editable=False,
                        hidden_columns=hidden_columns,
                    ),
                    style={"paddingTop": "16px"},
                ),
            )
        )
    return dcc.Tabs(tabs, className="artifact-tabs")


def build_editor_section(
    palette: dict[str, str],
    scenario: dict[str, Any],
    nodes_editor: pd.DataFrame,
    links_editor: pd.DataFrame,
) -> dbc.Accordion:
    editable = scenario["editable"]
    return dbc.Accordion(
        [
            dbc.AccordionItem(
                [
                    html.P(
                        "Activa o desactiva nodos para construir la topologia sobre la Meta-Red."
                        if editable
                        else "Referencia base de nodos de la Meta-Red.",
                        className="text-muted",
                    ),
                    build_dash_table(
                        nodes_editor,
                        palette,
                        table_id={"type": "scenario-nodes", "scenario": scenario["id"]},
                        table_kind="standard",
                        editable=editable,
                    ),
                ],
                title="Nodos",
            ),
            dbc.AccordionItem(
                [
                    html.P(
                        "Activa o desactiva enlaces concretos para cerrar o abrir rutas dentro del escenario."
                        if editable
                        else "Referencia base de enlaces de la Meta-Red.",
                        className="text-muted",
                    ),
                    build_dash_table(
                        links_editor,
                        palette,
                        table_id={"type": "scenario-links", "scenario": scenario["id"]},
                        table_kind="standard",
                        editable=editable,
                        hidden_columns=["row_id"],
                    ),
                ],
                title="Enlaces",
            ),
            dbc.AccordionItem(
                html.Div(id={"type": "scenario-artifacts", "scenario": scenario["id"]}),
                title="Matrices y metricas",
            ),
        ],
        start_collapsed=True,
        flush=True,
        always_open=False,
        className="mt-4",
    )


def build_topology_tab(
    base_model: dict[str, pd.DataFrame],
    palette: dict[str, str],
    scenario: dict[str, Any],
    nodes_editor: pd.DataFrame,
    links_editor: pd.DataFrame,
) -> dbc.Container:
    return dbc.Container(
        [
            html.P(scenario["description"], className="text-muted mb-3"),
            dbc.Row(
                [
                    dbc.Col(
                        dbc.Card(
                            dbc.CardBody(
                                cyto.Cytoscape(
                                    id={"type": "scenario-graph", "scenario": scenario["id"]},
                                    elements=[],
                                    stylesheet=build_cytoscape_stylesheet(palette),
                                    layout={"name": "preset", "fit": True, "padding": 40, "animate": False},
                                    style={"width": "100%", "height": "680px", "background": "#FFFFFF"},
                                    minZoom=0.35,
                                    maxZoom=1.8,
                                    wheelSensitivity=0.15,
                                )
                            ),
                            style={
                                "background": palette["card_bg"],
                                "border": f"1px solid {palette['card_border']}",
                                "borderRadius": "18px",
                                "boxShadow": "0 16px 40px rgba(31, 78, 120, 0.10)",
                            },
                        ),
                        xl=9,
                        lg=8,
                    ),
                    dbc.Col(
                        [
                            html.Div(id={"type": "scenario-graph-info", "scenario": scenario["id"]}),
                            html.Div(id={"type": "scenario-metrics", "scenario": scenario["id"]}),
                        ],
                        xl=3,
                        lg=4,
                    ),
                ],
                className="g-3",
            ),
            build_editor_section(palette, scenario, nodes_editor, links_editor),
        ],
        fluid=True,
        className="px-0",
    )


def build_comparison_placeholder(palette: dict[str, str]) -> dbc.Container:
    return dbc.Container(
        [
            html.P("Comparacion consolidada de metricas finales entre topologias.", className="text-muted mb-3"),
            dbc.Row(
                [
                    dbc.Col(
                        dcc.Dropdown(
                            id="comparison-metric-dropdown",
                            options=[{"label": metric, "value": metric} for metric in METRIC_COLUMN_ORDER],
                            value=METRIC_COLUMN_ORDER[0],
                            clearable=False,
                        ),
                        lg=4,
                    )
                ],
                className="mb-3",
            ),
            html.Div(id="comparison-table-wrapper"),
            dbc.Card(
                dbc.CardBody(dcc.Graph(id="comparison-chart", config={"displayModeBar": False})),
                style={
                    "background": palette["card_bg"],
                    "border": f"1px solid {palette['card_border']}",
                    "borderRadius": "18px",
                    "marginTop": "16px",
                },
            ),
        ],
        fluid=True,
        className="px-0",
    )


def build_dashboard_content(
    palette: dict[str, str],
    workbook_source: str,
) -> html.Div:
    topology_tabs = [
        dcc.Tab(label=scenario["label"], value=scenario["id"])
        for scenario in TOPOLOGY_DEFINITIONS
    ]
    topology_tabs.append(dcc.Tab(label="Comparacion", value="comparison"))

    return html.Div(
        [
            dbc.Alert(
                [
                    html.Span("Workbook cargado: ", style={"fontWeight": 700}),
                    html.Span(workbook_source),
                ],
                color="light",
                style={
                    "background": "#FFF8EC",
                    "border": f"1px solid {palette['card_border']}",
                    "color": palette["accent"],
                },
            ),
            dcc.Tabs(id="main-tabs", children=topology_tabs, value="meta_red"),
            html.Div(id="active-tab-content", style={"paddingTop": "20px"}),
        ]
    )


def build_error_alert(message: str) -> dbc.Alert:
    return dbc.Alert(message, color="danger", className="mt-4")


def build_initial_layout() -> dbc.Container:
    return dbc.Container(
        [
            dcc.Store(id="workbook-version-store", data=APP_STATE.get("version", 0)),
            dbc.Row(
                [
                    dbc.Col(
                        [
                            html.Div("Migracion Dash", style={"letterSpacing": "0.12em", "textTransform": "uppercase", "color": "#7B6A58", "fontSize": "0.78rem", "fontWeight": 700}),
                            html.H1(APP_TITLE, style={"fontWeight": 800, "marginBottom": "8px", "color": "#1F4E78"}),
                            html.P(
                                "Topologias FT con tablas estilizadas segun el workbook, grafos interactivos y comparacion consolidada.",
                                className="text-muted",
                            ),
                        ],
                        lg=8,
                    ),
                    dbc.Col(
                        dbc.Card(
                            dbc.CardBody(
                                [
                                    html.Div("Actualizar workbook", style={"fontWeight": 700, "marginBottom": "10px"}),
                                    dcc.Upload(
                                        id="workbook-upload",
                                        children=html.Div(["Arrastra un xlsm/xlsx o pulsa para cargarlo"]),
                                        multiple=False,
                                        style={
                                            "width": "100%",
                                            "padding": "18px",
                                            "border": "2px dashed #C9B9A1",
                                            "borderRadius": "14px",
                                            "textAlign": "center",
                                            "background": "#FFF8EC",
                                            "cursor": "pointer",
                                        },
                                    ),
                                ]
                            ),
                            style={
                                "background": "#FFFCF7",
                                "border": "1px solid #D7C9B6",
                                "borderRadius": "18px",
                            },
                        ),
                        lg=4,
                    ),
                ],
                className="align-items-center mb-4",
            ),
            html.Div(id="dashboard-root"),
        ],
        fluid=True,
        style={
            "minHeight": "100vh",
            "background": "linear-gradient(180deg, #F5F1E8 0%, #FBF8F2 48%, #EEF3F8 100%)",
            "padding": "24px 28px 40px 28px",
        },
    )


app = Dash(__name__, external_stylesheets=[dbc.themes.FLATLY], suppress_callback_exceptions=True)
app.title = APP_TITLE
initialize_app_state()
app.layout = build_initial_layout


@app.callback(
    Output("workbook-version-store", "data"),
    Input("workbook-upload", "contents"),
    State("workbook-upload", "filename"),
    prevent_initial_call=True,
)
def update_workbook_store(contents: str | None, filename: str | None):
    if not contents:
        raise PreventUpdate
    _, encoded = contents.split(",", 1)
    workbook_bytes = encoded.encode("ascii")
    import base64
    decoded = base64.b64decode(workbook_bytes)
    return set_app_workbook(decoded, filename or "workbook_subido")


@app.callback(
    Output("dashboard-root", "children"),
    Input("workbook-version-store", "data"),
)
def render_dashboard(_: int):
    if APP_STATE.get("base_model") is None or APP_STATE.get("palette") is None:
        return build_error_alert(APP_STATE.get("error") or "No hay workbook cargado.")

    try:
        return build_dashboard_content(
            dict(APP_STATE["palette"]),
            str(APP_STATE["workbook_source"]),
        )
    except WorkbookSchemaError as exc:
        return build_error_alert(f"El workbook no cumple el esquema esperado: {exc}")
    except KeyError as exc:
        return build_error_alert(f"Falta una tabla requerida en el workbook: {exc}")
    except Exception as exc:
        return build_error_alert(f"No se pudo cargar el dashboard Dash: {exc}")


@app.callback(
    Output("active-tab-content", "children"),
    Input("main-tabs", "value"),
    Input("workbook-version-store", "data"),
)
def render_active_tab(tab_value: str | None, _: int):
    if APP_STATE.get("base_model") is None or APP_STATE.get("palette") is None:
        return build_error_alert(APP_STATE.get("error") or "No hay workbook cargado.")

    if not tab_value:
        raise PreventUpdate

    palette = dict(APP_STATE["palette"])
    if tab_value == "comparison":
        return build_comparison_placeholder(palette)

    base_model = copy_frames(APP_STATE["base_model"])
    scenario = next((item for item in TOPOLOGY_DEFINITIONS if item["id"] == tab_value), None)
    if scenario is None:
        return build_error_alert(f"No existe la pestaña solicitada: {tab_value}")

    scenario_state = APP_STATE.get("scenario_states", {}).get(tab_value)
    if not scenario_state:
        scenario_state = initialize_scenario_states(base_model)[tab_value]
        APP_STATE.setdefault("scenario_states", {})[tab_value] = scenario_state

    nodes_editor = pd.DataFrame(scenario_state["nodes"])
    links_editor = pd.DataFrame(scenario_state["links"])
    return build_topology_tab(base_model, palette, scenario, nodes_editor, links_editor)


@app.callback(
    Output({"type": "scenario-graph", "scenario": MATCH}, "elements"),
    Output({"type": "scenario-metrics", "scenario": MATCH}, "children"),
    Output({"type": "scenario-artifacts", "scenario": MATCH}, "children"),
    Input({"type": "scenario-nodes", "scenario": MATCH}, "data"),
    Input({"type": "scenario-links", "scenario": MATCH}, "data"),
    State({"type": "scenario-nodes", "scenario": MATCH}, "id"),
)
def update_scenario_view(
    nodes_records: list[dict[str, Any]],
    links_records: list[dict[str, Any]],
    component_id: dict[str, str],
):
    if APP_STATE.get("base_model") is None or APP_STATE.get("palette") is None:
        raise PreventUpdate

    scenario_id = component_id["scenario"]
    base_model = copy_frames(APP_STATE["base_model"])
    palette = dict(APP_STATE["palette"])
    APP_STATE.setdefault("scenario_states", {})[scenario_id] = {
        "nodes": [dict(row) for row in nodes_records],
        "links": [dict(row) for row in links_records],
    }
    if scenario_id == "meta_red":
        scenario_model = copy_frames(base_model)
    else:
        scenario_model = apply_scenario_actives(base_model, nodes_records, links_records)

    artifacts = build_artifacts(scenario_model)
    elements = build_topology_graph_elements(scenario_model, artifacts["red"])
    metrics = build_metric_cards(artifacts["metricas_red"], palette)
    artifacts_tabs = build_artifact_tabs(artifacts, palette, scenario_id)
    return elements, metrics, artifacts_tabs


@app.callback(
    Output({"type": "scenario-graph-info", "scenario": MATCH}, "children"),
    Input({"type": "scenario-graph", "scenario": MATCH}, "tapNodeData"),
    Input({"type": "scenario-graph", "scenario": MATCH}, "tapEdgeData"),
)
def update_graph_info_panel(node_data: dict[str, Any] | None, edge_data: dict[str, Any] | None):
    if APP_STATE.get("palette") is None:
        raise PreventUpdate
    return build_graph_info_panel(dict(APP_STATE["palette"]), node_data=node_data, edge_data=edge_data)


@app.callback(
    Output("comparison-table-wrapper", "children"),
    Output("comparison-chart", "figure"),
    Input("main-tabs", "value"),
    Input("comparison-metric-dropdown", "value"),
)
def update_comparison_view(
    active_tab: str | None,
    selected_metric: str,
):
    if active_tab != "comparison":
        raise PreventUpdate
    if APP_STATE.get("base_model") is None or APP_STATE.get("palette") is None:
        raise PreventUpdate

    base_model = copy_frames(APP_STATE["base_model"])
    palette = dict(APP_STATE["palette"])
    rows = []
    for scenario in TOPOLOGY_DEFINITIONS:
        scenario_id = scenario["id"]
        scenario_state = APP_STATE.get("scenario_states", {}).get(scenario_id)
        if not scenario_state:
            continue
        if scenario_id == "meta_red":
            scenario_model = copy_frames(base_model)
        else:
            scenario_model = apply_scenario_actives(base_model, scenario_state["nodes"], scenario_state["links"])
        artifacts = build_artifacts(scenario_model)
        metric_values = artifacts["metricas_red"][["metrica", "valor"]].copy()
        metric_values["valor"] = pd.to_numeric(metric_values["valor"], errors="coerce").fillna(0.0)
        metric_map = dict(zip(metric_values["metrica"], metric_values["valor"]))
        row = {
            "Topologia": scenario["label"],
            "Nodos activos": int((scenario_model["nodos"]["activo"] == 1).sum()),
            "Enlaces activos": int((artifacts["red"]["activo"] == 1).sum()),
        }
        for metric_name in METRIC_COLUMN_ORDER:
            row[metric_name] = float(metric_map.get(metric_name, 0.0))
        rows.append(row)

    comparison = pd.DataFrame(rows)
    table = build_dash_table(
        comparison,
        palette,
        table_id="comparison-table",
        table_kind="metrics",
        editable=False,
    )
    figure = px.bar(
        comparison,
        x="Topologia",
        y=selected_metric,
        color="Topologia",
        text_auto=".4f",
        color_discrete_sequence=["#1F4E78", "#D66D75", "#E8A66B", "#5DA9E9"],
    )
    figure.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin={"l": 20, "r": 20, "t": 30, "b": 20},
        legend_title_text="",
        yaxis_title=selected_metric,
        xaxis_title="",
    )
    return table, figure


if __name__ == "__main__":
    app.run(debug=True)